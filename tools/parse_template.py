# -*- coding: utf-8 -*-
"""Разбор «Шаблон_ГРП v2.xlsx» в машинный вид + сверка статистики typGRP.md §3.

Выход:
  tests/template_parsed.json  — 1 672 задачи в порядке обхода дерева
  печать отчёта сверки в stdout

Отличие от версии для шаблона v1: колонки «СДР» больше нет, «Уровень структуры»
подаётся самим шаблоном (typGRP.md §2.2) и не вычисляется. Колонка
«Тип ограничения» сохраняет разрешённые ограничения Microsoft Project.
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl

# Консоль Windows может быть в cp1251: не даём выводу падать на символах,
# которых нет в её кодировке (стрелки, типографика).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "instructions" / "Шаблон_ГРП v2.xlsx"
DST = ROOT / "tests" / "template_parsed.json"

SOURCE_COLUMNS = [
    "Тип ограничения", "Вид работ", "Код классификатора", "Уровень структуры", "Ид.",
    "Название задачи", "% завершения", "Длительность", "Начало", "Окончание",
    "Критическая задача", "Предшественники", "Последователи", "комментарий",
]
COLUMNS = [name for name in SOURCE_COLUMNS if name != "Критическая задача"]

# Заявлено в typGRP.md §3 — сверяем, а не доверяем.
DECLARED = {
    "всего задач": 1673,
    "вех": 293,
    "со связями": 1220,
    "с комментариями": 260,
    "суммарных строк": 420,
    "уровни": {1: 18, 2: 84, 3: 280, 4: 636, 5: 336, 6: 174, 7: 88, 8: 57},
}


def cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load() -> list[dict]:
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [cell(c) for c in next(rows)]
    if header != SOURCE_COLUMNS:
        raise SystemExit(
            f"Шапка шаблона изменилась.\nОжидалось: {SOURCE_COLUMNS}\nПолучено:  {header}")

    tasks = []
    for raw in rows:
        source = {name: cell(v) for name, v in zip(SOURCE_COLUMNS, raw)}
        rec = {name: source[name] for name in COLUMNS}
        if not rec["Ид."] and not rec["Название задачи"]:
            continue
        rec["Уровень структуры"] = int(rec["Уровень структуры"])
        rec["Название задачи"] = rec["Название задачи"].strip()
        tasks.append(rec)
    return tasks


def summary_ids(tasks: list[dict]) -> set[str]:
    """Суммарная строка — та, за которой непосредственно следует строка большего уровня."""
    out = set()
    for i, t in enumerate(tasks[:-1]):
        if tasks[i + 1]["Уровень структуры"] > t["Уровень структуры"]:
            out.add(t["Ид."])
    return out


def report(tasks: list[dict]) -> int:
    levels = Counter(t["Уровень структуры"] for t in tasks)
    summ = summary_ids(tasks)
    facts = {
        "всего задач": len(tasks),
        "вех": sum(1 for t in tasks if t["Длительность"].startswith("0 дн")),
        "со связями": sum(1 for t in tasks if t["Предшественники"]),
        "с комментариями": sum(1 for t in tasks if t["комментарий"]),
        "суммарных строк": len(summ),
        "уровни": dict(sorted(levels.items())),
    }

    print("=== Сверка со статистикой typGRP.md §3 ===")
    bad = 0
    for key in ("всего задач", "вех", "со связями", "с комментариями", "суммарных строк"):
        ok = facts[key] == DECLARED[key]
        bad += 0 if ok else 1
        print(f"{'OK  ' if ok else 'РАСХ'} {key}: факт {facts[key]}, заявлено {DECLARED[key]}")

    ok_levels = facts["уровни"] == DECLARED["уровни"]
    bad += 0 if ok_levels else 1
    print(f"{'OK  ' if ok_levels else 'РАСХ'} распределение по уровням: факт {facts['уровни']}")
    if not ok_levels:
        print(f"     заявлено {DECLARED['уровни']}")

    # Проверки, важные для импорта в MS Project (typGRP.md §2.2)
    print("\n=== Проверки импорта (typGRP.md §2.2) ===")
    first = tasks[0]["Уровень структуры"]
    print(f"{'OK  ' if first == 1 else 'РАСХ'} уровень первой строки: {first}")

    jumps = [
        (i, tasks[i - 1]["Название задачи"], tasks[i]["Название задачи"])
        for i in range(1, len(tasks))
        if tasks[i]["Уровень структуры"] - tasks[i - 1]["Уровень структуры"] > 1
    ]
    print(f"{'OK  ' if not jumps else 'РАСХ'} скачков уровня >1: {len(jumps)}")
    for i, prev, cur in jumps[:10]:
        print(f"     строка {i + 2}: {prev[:40]} -> {cur[:40]}")

    ids = [int(t["Ид."]) for t in tasks]
    gaps = sorted(set(range(1, max(ids) + 1)) - set(ids))
    print(f"{'OK  ' if not gaps else 'РАСХ'} Ид. сквозной от 1 без разрывов; "
          f"пропущено {len(gaps)}: {gaps}")
    print("     (typGRP.md §13 расх. 1: нумерация шаблона в выдачу не переносится)")

    known = set(t["Ид."] for t in tasks)
    broken = [
        (t["Ид."], t["Название задачи"][:45], p)
        for t in tasks
        for p in (x.strip() for x in t["Предшественники"].split(";") if x.strip())
        if (m := re.match(r"^(\d+)", p)) and m.group(1) not in known
    ]
    print(f"{'OK  ' if not broken else 'РАСХ'} битых ссылок в «Предшественниках»: {len(broken)}")
    for b in broken:
        print(f"     Ид. {b[0]} «{b[1]}» -> {b[2]}")
    print("     (typGRP.md §13 расх. 2: восстанавливаются на шаги блока «Ограждение территории»)")

    to_summary = sum(
        1
        for t in tasks
        for p in (x.strip() for x in t["Предшественники"].split(";") if x.strip())
        if (m := re.match(r"^(\d+)", p)) and m.group(1) in summ
    )
    print(f"ИНФО связей на суммарные строки: {to_summary} (§13 расх. 6 — сохраняются как есть)")

    succ_with_code = sum(
        1 for t in tasks
        if t["Последователи"] and re.search(r"\d+(ОН|НН|ОО|НО)|[+-]\d+\s*дн", t["Последователи"])
    )
    print(f"ИНФО «Последователей» с кодом связи и лагом: {succ_with_code} (§13 расх. 8 — код отбрасывается)")

    dirty = [t["Ид."] for t in tasks if t["Ид."] in summ and (t["Длительность"] or t["% завершения"])]
    print(f"ИНФО суммарных строк с собственной длительностью/процентом: {len(dirty)}")
    print("     (в шаблоне это норма; в выдаче агента такие поля пусты — §2.2 п. 4)")

    mdm = sum(1 for t in tasks if t["Вид работ"] or t["Код классификатора"])
    print(f"ИНФО строк с заполненным МДМ («Вид работ» / «Код классификатора»): {mdm}")
    print("     (§13 расх. 14: пока пусты, выводятся пустыми со строкой в «Допущения»)")

    lat = sum(1 for t in tasks if re.search(r"\d+(FS|SS|FF|SF)", t["Предшественники"]))
    print(f"{'OK  ' if not lat else 'РАСХ'} латинских кодов связей: {lat} (допустимы только ОН/НН/ОО/НО)")

    return bad


def main() -> int:
    tasks = load()
    DST.parent.mkdir(exist_ok=True)
    DST.write_text(
        json.dumps(tasks, ensure_ascii=False, indent=1),
        encoding="utf-8",
    )
    print(f"Записано: {DST.relative_to(ROOT)} ({len(tasks)} задач)\n")
    return 1 if report(tasks) else 0


if __name__ == "__main__":
    sys.exit(main())
