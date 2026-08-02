# -*- coding: utf-8 -*-
"""Валидатор выдачи ГРП по машинно проверяемой части чек-листа CLAUDE.md §9.

Использование:
    python tools/validate_grp.py <файл.xlsx>
    python tools/validate_grp.py tests/template_parsed.json     # самопроверка на шаблоне

Проверяет только то, что проверяется формально. Пункты чек-листа, требующие
предметного суждения (сезонный гейт, здравый смысл критического пути, оценка
уверенности), остаются за планировщиком и здесь НЕ подменяются.

Формат — шаблон v2: 12 колонок, колонки СДР нет. Иерархия читается только из
«Уровня структуры»; адрес строки в сообщениях — «Ид. + наименование».

Код возврата: 0 — нарушений уровня ОШИБКА нет; 1 — есть.
"""
from __future__ import annotations

import json
import re
import sys
from datetime import datetime
from pathlib import Path

# Консоль Windows может быть в cp1251: не даём выводу падать на символах,
# которых нет в её кодировке (стрелки, типографика).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")

ROOT = Path(__file__).resolve().parents[1]

LATIN_LINK = re.compile(r"\d+\s*(FS|SS|FF|SF)\b", re.IGNORECASE)
LINK_RE = re.compile(r"^(\d+)\s*([А-Я]{2})?\s*([+-]\s*\d+)?")
VAGUE = re.compile(r"(≥|≤|не\s+менее|не\s+более|более|около|примерно|~)")
DUR_RE = re.compile(r"^\d+([.,]\d+)?\s*(дн|день|дня|дней)\.?\??$", re.IGNORECASE)
SUCC_DIRTY = re.compile(r"(ОН|НН|ОО|НО)|дн")

# CLAUDE.md §9 «Все обязательные вехи присутствуют»
REQUIRED_MILESTONES = {
    "АК (архитектурный конкурс)": ("архитектурн", "конкурс"),
    "ГПЗУ": ("гпзу",),
    "РС": ("разрешени", "строительств"),
    "Завершение монолита": ("монолит",),
    "Замыкание теплового контура": ("теплов", "контур"),
    "Пуск тепла": ("пуск", "тепл"),
    "ЗОС": ("зос",),
    "РВЭ": ("рвэ",),
    "Передача": ("передан",),
}

# typGRP.md §2 — формат строки шаблона v2
COLUMNS = ["Вид работ", "Код классификатора", "Уровень структуры", "Ид.",
           "Название задачи", "% завершения", "Длительность", "Начало", "Окончание",
           "Предшественники", "Последователи", "комментарий"]


class Report:
    def __init__(self) -> None:
        self.errors: list[str] = []
        self.warns: list[str] = []
        self.oks: list[str] = []

    def check(self, cond: bool, name: str, detail: str = "", warn_only: bool = False) -> None:
        if cond:
            self.oks.append(name)
        elif warn_only:
            self.warns.append(f"{name}{': ' + detail if detail else ''}")
        else:
            self.errors.append(f"{name}{': ' + detail if detail else ''}")

    def dump(self) -> int:
        for n in self.oks:
            print(f"  OK      {n}")
        for n in self.warns:
            print(f"  ВНИМАНИЕ {n}")
        for n in self.errors:
            print(f"  ОШИБКА  {n}")
        print(f"\nИтог: OK {len(self.oks)} · внимание {len(self.warns)} · "
              f"ошибок {len(self.errors)}")
        return 1 if self.errors else 0


def cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def load(path: Path) -> list[dict]:
    if path.suffix.lower() == ".json":
        return [{k: (v if isinstance(v, str) else str(v)) for k, v in r.items()}
                for r in json.loads(path.read_text(encoding="utf-8"))]

    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [cell(c) for c in next(rows)]
    tasks = []
    for raw in rows:
        rec = {h: cell(v) for h, v in zip(header, raw)}
        if not rec.get("Название задачи"):
            continue
        tasks.append(rec)
    return tasks


def label(t: dict) -> str:
    """Адрес строки для сообщений. СДР больше нет — «Ид. + наименование»."""
    return f"{t.get('Ид.', '?')} «{t.get('Название задачи', '')[:38]}»"


def validate(tasks: list[dict]) -> int:
    r = Report()
    n = len(tasks)
    print(f"Задач: {n}\n")

    # --- Формат строки (typGRP.md §2) -------------------------------------
    missing_cols = [c for c in COLUMNS if c not in tasks[0]]
    r.check(not missing_cols, "выведены все 12 колонок формата v2", f"нет: {missing_cols}")
    if "Уровень структуры" not in tasks[0]:
        r.errors.append("колонка «Уровень структуры» отсутствует — иерархию восстановить "
                        "не из чего (СДР в шаблоне v2 нет, typGRP.md §2.2)")
        return r.dump()
    for t in tasks:
        t["_lvl"] = int(float(t["Уровень структуры"])) if t.get("Уровень структуры") else 0

    if any(t.get("Вид работ") or t.get("Код классификатора") for t in tasks):
        r.warns.append("справочно: часть строк несёт коды МДМ — справочник поступил, "
                       "проверьте соответствие классификатору")
    else:
        r.oks.append("колонки МДМ пусты — справочник не поступил (typGRP.md §13 расх. 14)")

    # --- Иерархия (typGRP.md §2.2) ---------------------------------------
    r.check(tasks[0]["_lvl"] == 1, "уровень первой строки = 1", f"фактически {tasks[0]['_lvl']}")

    jumps = [(i + 2, label(tasks[i])) for i in range(1, n)
             if tasks[i]["_lvl"] - tasks[i - 1]["_lvl"] > 1]
    r.check(not jumps, "уровень не растёт больше чем на 1",
            f"{len(jumps)} скачков, напр. {jumps[:3]}")

    bad_lvl = [label(t) for t in tasks if t["_lvl"] < 1]
    r.check(not bad_lvl, "уровень структуры заполнен у всех строк",
            f"{len(bad_lvl)} пустых, напр. {bad_lvl[:5]}")

    # Суммарная строка — та, за которой следует строка большего уровня.
    summary_idx = {i for i in range(n - 1) if tasks[i + 1]["_lvl"] > tasks[i]["_lvl"]}

    # --- Суммарные строки -------------------------------------------------
    # §2.2 п.4: у суммарных строк пусты ДЛИТЕЛЬНОСТЬ и % ЗАВЕРШЕНИЯ — именно они
    # ломают импорт. Даты выводятся (DEC-25), связи сохраняются (DEC-26).
    dirty = [label(tasks[i]) for i in summary_idx
             if tasks[i].get("Длительность") or tasks[i].get("% завершения")]
    r.check(not dirty, "у суммарных строк пусты длительность и % завершения",
            f"{len(dirty)} нарушений (§2.2 п.4), напр. {dirty[:5]}")

    undated = [label(tasks[i]) for i in summary_idx
               if not (tasks[i].get("Начало") and tasks[i].get("Окончание"))]
    r.check(not undated, "у суммарных строк проставлены даты (DEC-25)",
            f"{len(undated)} без дат, напр. {undated[:5]}", warn_only=True)

    # --- Ид. ---------------------------------------------------------------
    ids = [t.get("Ид.", "") for t in tasks]
    nums = [int(x) for x in ids if x.isdigit()]
    r.check(len(nums) == n, "все Ид. заполнены и числовые", f"{n - len(nums)} пустых/нечисловых")
    if len(nums) == n:
        gaps = sorted(set(range(min(nums), max(nums) + 1)) - set(nums))
        r.check(not gaps, "столбец Ид. сквозной, без разрывов",
                f"{len(gaps)} пропущено: {gaps[:12]}")
        r.check(len(set(nums)) == len(nums), "Ид. уникальны",
                f"{len(nums) - len(set(nums))} дубликатов")
        r.check(nums == list(range(1, n + 1)), "Ид. совпадает с номером строки",
                "нумерация сдвинута — связи после импорта укажут на чужие задачи")

    present_ids = set(nums)

    # --- Связи -------------------------------------------------------------
    latin = [label(t) for t in tasks if LATIN_LINK.search(t.get("Предшественники", ""))]
    r.check(not latin, "нотация связей русская (ОН · НН · ОО · НО)",
            f"латиница в {len(latin)} строках, напр. {latin[:5]}")

    vague = [label(t) for t in tasks if VAGUE.search(t.get("Предшественники", ""))]
    r.check(not vague, "лаги численные, без формулировок «не менее»",
            f"{len(vague)} строк, напр. {vague[:5]}")

    # DEC-27: «Последователи» — только Ид. через «;», без кода связи и лага.
    succ_dirty = [label(t) for t in tasks if SUCC_DIRTY.search(t.get("Последователи", ""))]
    r.check(not succ_dirty, "«Последователи» — только Ид., без кода связи и лага (DEC-27)",
            f"{len(succ_dirty)} строк, напр. {succ_dirty[:5]}")

    dangling, to_summary = [], []
    idx_by_id = {int(t["Ид."]): i for i, t in enumerate(tasks) if t.get("Ид.", "").isdigit()}
    has_successor = set()
    for t in tasks:
        for part in (x.strip() for x in t.get("Предшественники", "").split(";")):
            if not part:
                continue
            m = LINK_RE.match(part)
            if not m:
                continue
            pid = int(m.group(1))
            if pid not in present_ids:
                dangling.append((label(t), part))
            else:
                has_successor.add(pid)
                if idx_by_id.get(pid) in summary_idx:
                    to_summary.append((label(t), part))

    r.check(not dangling, "все связи ведут на существующие Ид.",
            f"{len(dangling)} битых, напр. {dangling[:5]}")
    # Связь на суммарную строку — норма, а не нарушение (typGRP.md §2.2, DEC-26).
    if to_summary:
        r.warns.append(f"справочно: {len(to_summary)} связей ведут на суммарные строки — "
                       f"это норма (typGRP.md §2.2), напр. {to_summary[:3]}")
    else:
        r.oks.append("связей на суммарные строки нет")

    leaves = [(i, t) for i, t in enumerate(tasks) if i not in summary_idx]
    no_pred = [label(t) for _, t in leaves if not t.get("Предшественники")]
    no_succ = [label(t) for _, t in leaves
               if t.get("Ид.", "").isdigit() and int(t["Ид."]) not in has_successor
               and not t.get("Последователи")]
    r.check(len(no_pred) <= 1, "нет задач-листьев без предшественников (кроме стартовой)",
            f"{len(no_pred)} шт., напр. {no_pred[:6]}", warn_only=True)
    r.check(len(no_succ) <= 1, "нет задач-листьев без последователей (кроме финальной)",
            f"{len(no_succ)} шт., напр. {no_succ[:6]}", warn_only=True)

    # --- Длительности ------------------------------------------------------
    bad_dur = [(label(t), t["Длительность"]) for _, t in leaves
               if t.get("Длительность") and not DUR_RE.match(t["Длительность"])]
    r.check(not bad_dur, "длительности в формате «N дней»",
            f"{len(bad_dur)} нарушений, напр. {bad_dur[:5]}")

    prelim = [label(t) for t in tasks if t.get("Длительность", "").endswith("?")]
    r.check(not prelim, "нет предварительных оценок (суффикс «?»)",
            f"{len(prelim)} шт., напр. {prelim[:5]}", warn_only=True)

    # --- Обязательные вехи -------------------------------------------------
    milestones = [t for t in tasks if t.get("Длительность", "").startswith("0 дн")]
    ms_names = [t["Название задачи"].lower() for t in milestones]
    missing = [name for name, keys in REQUIRED_MILESTONES.items()
               if not any(all(k in nm for k in keys) for nm in ms_names)]
    r.check(not missing, "все обязательные вехи присутствуют", f"нет: {missing}")

    # --- DEC-30: согласованность ветки чистовой отделки квартир -------------
    # Список намеренно дублирует MARKERS из tests/test_dec30.py — см. план,
    # задача 3: тест и валидатор должны падать независимо друг от друга.
    FLAT_FIT_MARKERS = (
        "отделочные работы квартиры",
        "подготовка под чистовую отделку квартир",
        "чистовая отделка квартир",
        "тендер отделка квартиры чистовая",
        "мокап отделки типового этажа",
        "завершены отделочные работы квартир",
        "передача квартир с отделкой",
        "переданы квартиры с отделкой",
        "отделка квартиры",
    )
    lowered = [t["Название задачи"].lower() for t in tasks]
    present = [m for m in FLAT_FIT_MARKERS if any(m in nm for nm in lowered)]
    r.check(len(present) in (0, len(FLAT_FIT_MARKERS)),
            "ветка чистовой отделки квартир согласована (DEC-30)",
            f"снята частично: осталось {len(present)} из {len(FLAT_FIT_MARKERS)} маркеров — "
            f"{present}")

    # --- Передача квартир без отделки = РВЭ + 180 дн (standards.md §15.3) ---
    rve = next((t for t in tasks
                if t["Название задачи"].strip().lower().startswith("рвэ по этапу")), None)
    handovers = [t for t in tasks
                 if t["Название задачи"].strip().lower() == "передача квартир без отделки"]
    hand = min(handovers, key=lambda t: t["_lvl"]) if handovers else None
    if rve and hand and rve.get("Окончание") and hand.get("Начало"):
        delta = (datetime.strptime(hand["Начало"], "%d.%m.%Y")
                 - datetime.strptime(rve["Окончание"], "%d.%m.%Y")).days
        r.check(delta == 180,
                "передача квартир без отделки = РВЭ + 180 дн (standards.md §15.3)",
                f"фактически +{delta} дн")
    else:
        r.warns.append("справочно: веха «Передача квартир без отделки» или «РВЭ по этапу» "
                       "не найдена — проверка срока передачи пропущена")

    return r.dump()


def main() -> int:
    if len(sys.argv) != 2:
        print(__doc__)
        return 2
    path = Path(sys.argv[1])
    if not path.is_absolute():
        path = ROOT / path
    if not path.exists():
        print(f"Файл не найден: {path}")
        return 2
    print(f"=== Валидация: {path.name} (CLAUDE.md §9, машинная часть) ===\n")
    return validate(load(path))


if __name__ == "__main__":
    sys.exit(main())
