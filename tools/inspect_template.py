# -*- coding: utf-8 -*-
"""Разведка структуры Шаблон_ГРП.xlsx: листы, шапка, первые строки."""
import sys
from pathlib import Path

import openpyxl

# Консоль Windows может быть в cp1251: не даём выводу падать на символах,
# которых нет в её кодировке (стрелки, типографика).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "instructions" / "Шаблон_ГРП.xlsx"


def main() -> int:
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    print("Листы:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n=== {name}: {ws.max_row} строк x {ws.max_column} колонок ===")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
            cells = ["" if c is None else str(c) for c in row]
            print(f"{i}: {' | '.join(cells)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
