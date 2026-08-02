# -*- coding: utf-8 -*-
"""Генератор ГРП: развёртывание каркаса typGRP.md v4.0 в Excel для импорта в MS Project.

    python tools/build_grp.py tests/etalon_project.json out/ГРП.xlsx

Конвейер:
    1. каркас из шаблона v2 (tests/template_parsed.json)
    2. устранение расхождений typGRP.md §13 (расх. 1–3)
    3. наследование связей суммарных строк листьями
    4. условия площадки — typGRP.md §6.8
    5. состав объектов: паркинг и корпуса по ТЭП (два прототипа по остеклению)
    6. нормативы standards.md, поэтажная развёртка монолита
    7. сетевой расчёт (CPM), проход 1
    8. тепловой контур и отделка по каждому корпусу (bindings.md §3.6, §3.8)
    9. сетевой расчёт, проход 2
   10. перенумерация Ид., очистка суммарных строк, запись Excel

Адресация. Колонки СДР в шаблоне v2 нет (typGRP.md §2), поэтому:
    · внутренний ключ строки — «Ид.» шаблона; вставленные строки получают «n<N>»
    · иерархия — только «Уровень структуры»: родитель есть ближайшая строка выше
      с уровнем на единицу меньше
    · правила bindings.md адресуются парой «префикс объекта + наименование»

Что генератор НЕ делает (по решениям владельца, а не по недосмотру):
    · сценарии и вероятностный расчёт — DEC-16, правила не заданы
    · ресурсный контур — CLAUDE.md §4
    · два и более этапа ввода — правил разнесения объектов нет, typGRP.md §14 п.7
    · утверждение графика — CLAUDE.md §1
"""
from __future__ import annotations

import json
import re
import sys
from datetime import date, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from grp_model import (  # noqa: E402
    Assumption, Bnd, Node, Std, backward_pass, compute_thermal_and_fit, dfmt, dparse,
    fmt_links, forward_pass, monolith_floor_durations, parse_links, pile_duration,
    seasonal_duration,
)

# Консоль Windows может быть в cp1251: не даём выводу падать на символах,
# которых нет в её кодировке (стрелки, типографика).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")

ROOT = Path(__file__).resolve().parents[1]
SKELETON = ROOT / "tests" / "template_parsed.json"

COLUMNS = ["Вид работ", "Код классификатора", "Уровень структуры", "Ид.",
           "Название задачи", "% завершения", "Длительность", "Начало", "Окончание",
           "Предшественники", "Последователи", "комментарий"]

# Старт эталонного проекта — typGRP.md §1.1. База для переноса календарных якорей.
ETALON_START = dparse("10.01.2023")

PREFIX_RE = re.compile(r"^([А-ЯЁ]+\d+)\.\s*")


def strip_prefix(name: str) -> tuple[str, str]:
    """«К1. Монтаж фасадов» → («К1», «Монтаж фасадов»)."""
    m = PREFIX_RE.match(name)
    return (m.group(1), name[m.end():]) if m else ("", name)


class Build:
    def __init__(self, project: dict) -> None:
        self.p = project
        self.assumptions: list[Assumption] = []
        self.rationale: list[dict] = []
        self.notices: list[str] = []
        self.rows: list[dict] = []
        self.start = dparse(project["старт_проекта"])
        # Календарные якоря шаблона — абсолютные даты его собственного проекта.
        # Для любого другого проекта они переносятся на разницу стартов, иначе
        # регламентные вехи фазы A встают раньше начала работ.
        self.shift = (self.start - ETALON_START).days
        self._seq = 0
        self.corpus_prefix: dict[str, str] = {}     # код корпуса → префикс в графике
        self.fit_tasks: list[tuple[str, str]] = []  # (ключ задачи отделки, код объекта)
        self.parking_prefix: str | None = None

    # -- служебное ------------------------------------------------------
    def note(self, code: str, text: str, scope: str = "") -> None:
        self.assumptions.append(Assumption(code, text, scope))

    def why(self, block: str, value: str, source: str, confidence: str, comment: str = "") -> None:
        self.rationale.append({"Блок": block, "Значение": value, "Источник": source,
                               "Уверенность": confidence, "Комментарий": comment})

    def newkey(self) -> str:
        self._seq += 1
        return f"n{self._seq}"

    def by_key(self) -> dict[str, dict]:
        return {r["key"]: r for r in self.rows}

    # -- навигация по уровням (СДР больше нет) ---------------------------
    def index_of(self, key: str) -> int:
        for i, r in enumerate(self.rows):
            if r["key"] == key:
                return i
        raise KeyError(key)

    def subtree(self, i: int) -> list[int]:
        """Индексы всех потомков строки i (уровень строго больше)."""
        lvl = self.rows[i]["lvl"]
        out = []
        for j in range(i + 1, len(self.rows)):
            if self.rows[j]["lvl"] <= lvl:
                break
            out.append(j)
        return out

    def is_summary(self, i: int) -> bool:
        return i + 1 < len(self.rows) and self.rows[i + 1]["lvl"] > self.rows[i]["lvl"]

    def summaries(self) -> set[str]:
        return {self.rows[i]["key"] for i in range(len(self.rows)) if self.is_summary(i)}

    def ancestors(self, i: int) -> list[int]:
        """Индексы предков строки i, от ближайшего к корню."""
        out, lvl = [], self.rows[i]["lvl"]
        for j in range(i - 1, -1, -1):
            if self.rows[j]["lvl"] < lvl:
                out.append(j)
                lvl = self.rows[j]["lvl"]
                if lvl == 1:
                    break
        return out

    def find(self, name: str, prefix: str | None = None, exact: bool = True) -> list[int]:
        """Поиск строк по наименованию (и, при необходимости, префиксу объекта)."""
        out = []
        for i, r in enumerate(self.rows):
            pref, bare = strip_prefix(r["name"])
            if prefix is not None and pref != prefix:
                continue
            if (bare == name) if exact else bare.startswith(name):
                out.append(i)
        return out

    def one(self, name: str, prefix: str | None = None, exact: bool = True) -> int | None:
        hits = self.find(name, prefix, exact)
        return hits[0] if hits else None

    # ==================================================================
    # 1. Каркас
    # ==================================================================
    def load_skeleton(self) -> None:
        raw = json.loads(SKELETON.read_text(encoding="utf-8"))
        known = {r["Ид."] for r in raw}
        self.rows = []
        for r in raw:
            links = [(str(pid), kind, lag)
                     for pid, kind, lag in parse_links(r["Предшественники"])
                     if str(pid) in known]
            self.rows.append({
                "key": r["Ид."],
                "lvl": r["Уровень структуры"],
                "name": r["Название задачи"],
                "dur": self._dur(r["Длительность"]),
                "links": links,
                "comment": r["комментарий"],
                "tpl_start": r["Начало"],
                "src": "шаблон v2 (typGRP.md)",
            })
        self.why("Каркас", f"{len(self.rows)} строк",
                 "typGRP.md §3, шаблон v2 — 1 672 задачи", "высокая",
                 "Регламентные блоки фазы A наследуют длительности шаблона — typGRP.md §6.0")

    @staticmethod
    def _dur(s: str) -> int | None:
        if not s:
            return None
        m = re.match(r"^\s*(\d+)", s.replace(",", "."))
        return int(m.group(1)) if m else None

    # ==================================================================
    # 2. Расхождения шаблона — typGRP.md §13, расх. 1–3
    # ==================================================================
    NOMINATION_SHORT = [("РД ВПР номинация", 0), ("Подготовка ТЗ и ключевых условия номинация", 7),
                        ("Подготовка оферты номинация", 7), ("Тендер номинация", 7),
                        ("Выбор поставщика/подписание соглашения о намерениях номинация", 7)]
    TENDER_MAIN = [("РД ВПР", 0), ("Подготовка ТЗ", 7), ("Подготовка ВОР", 4),
                   ("Подготовка оферты", 3), ("Согласование тендерного пакета", 3),
                   ("Тендер", 30), ("Заключение договора", 14)]

    def repair_defects(self) -> None:
        def expand(head_name: str, steps: list[tuple[str, int]], subject: str) -> list[str]:
            i = self.one(head_name)
            if i is None:
                return []
            head = self.rows[i]
            head["dur"] = None            # суммарная строка — §2.2 п. 4
            head_links, head["links"] = head["links"], []
            new, prev = [], None
            for label, dur in steps:
                k = self.newkey()
                new.append({"key": k, "lvl": head["lvl"] + 1,
                            "name": f"{label} {subject}", "dur": dur,
                            "links": [(prev, "ОН", 0)] if prev else list(head_links),
                            "comment": "", "tpl_start": "",
                            "src": "восстановлено, typGRP.md §13 расх. 3"})
                prev = k
            self.rows[i + 1:i + 1] = new
            return [r["key"] for r in new]

        # Шаблон от 02.08.2026 закрыл блок «Ограждение территории» и вместе с ним
        # три битые связи на вехах МАФ. Пустым остался один блок номинации.
        expand("Тендер номинация Арматурный каркас сваи", self.NOMINATION_SHORT,
               "Арматурный каркас сваи")

        self.note("typGRP.md §13 расх. 1",
                  "Нумерация Ид. шаблона содержала разрыв (640–644) и в выдачу не перенесена: "
                  "Ид. присвоены заново сквозным рядом, все ссылки пересчитаны.")
        self.note("typGRP.md §13 расх. 3",
                  "Пустой блок «Тендер номинация Арматурный каркас сваи» развёрнут по канону "
                  "§9.1а (короткая номинация, 28 дн). Требует подтверждения владельца.")
        self.note("typGRP.md §13 расх. 14",
                  "Колонки «Вид работ» и «Код классификатора» выведены пустыми: справочник МДМ "
                  "не поступил. Собственные коды не изобретаются (CLAUDE.md §12).")

    # ==================================================================
    # 3. Наследование связей суммарных строк — решение владельца 01.08.2026
    # ==================================================================
    def inherit_summary_links(self) -> None:
        """Лист без предшественников берёт связи ближайшего предка-суммарной строки.

        В шаблоне v2 часть блоков ВИС несёт связи на суммарной строке, а её
        потомки-листья предшественников не имеют вовсе (например «П1. Внутренние
        сантехнические системы» → «П1. Хозяйственно-питьевой водопровод» → лист
        «По договору …»). При прямом переносе такие листья встают на старт проекта
        и утягивают расчёт. Сама связь на суммарной строке сохраняется — DEC-26.
        """
        inherited = 0
        for i, r in enumerate(self.rows):
            if self.is_summary(i) or r["links"]:
                continue
            for j in self.ancestors(i):
                if self.rows[j]["links"]:
                    r["links"] = list(self.rows[j]["links"])
                    r["src"] = f"связи унаследованы от «{self.rows[j]['name'][:40]}»"
                    inherited += 1
                    break
        if inherited:
            self.note("DEC-26 / решение 01.08.2026",
                      f"{inherited} задач-листьев не имели собственных предшественников, тогда как "
                      f"их суммарная строка-предок связи несёт. Листья унаследовали связи предка; "
                      f"на суммарной строке связь сохранена как в шаблоне. Без этого листья "
                      f"встали бы на старт проекта.")
            self.why("Наследование связей", f"{inherited} листьев",
                     "typGRP.md §2.2 · DEC-26", "средняя",
                     "Шаблон v2 непоследователен: часть блоков ВИС несёт связи только на "
                     "суммарной строке")

    # ==================================================================
    # 4. Условия площадки — typGRP.md §6.8
    # ==================================================================
    OPTIONAL_A = (
        ("Смена ВРИ",                          "ВРИ",            "Смена ВРИ"),
        ("Разработка и согласование ППТ",      "ППТ",            "ППТ и корректировка ПЗЗ"),
        ("Археологические изыскания",          "ОКН",            "Археологические изыскания"),
        ("Разработка ПД по СЗЗ (при необходимости)", "СЗЗ",      "Разработка ПД по СЗЗ"),
        ("Документация на снос",               "снос_застройки", "Документация на снос"),
        ("Снос зданий и сооружений",           "снос_застройки", "Снос зданий и сооружений"),
        ("Вынос сетей из пятна застройки",     "вынос_сетей",    "Вынос сетей из пятна застройки"),
    )

    def apply_site_conditions(self) -> None:
        """Опциональный блок включается только по признаку ТЭП.

        Неподтверждённое условие делает задачу неприменимой (CLAUDE.md §2.3), но не
        несуществующей: строка остаётся в структуре — состав разделов есть контракт
        с Агентом 4 — и обнуляется, чтобы не удерживать даты. Удаление строки порвало
        бы WBS и нарушило typGRP.md §2.2 п. 2.
        """
        cond = self.p.get("индивидуальные_условия", {})
        for name, key, label in self.OPTIONAL_A:
            if cond.get(key, False):
                continue
            i = self.one(name)
            if i is None:
                continue
            zeroed = 0
            for j in [i] + self.subtree(i):
                if self.rows[j]["dur"]:
                    self.rows[j]["dur"] = 0
                    self.rows[j]["src"] = f"typGRP.md §6.8 — условие «{key}» не подтверждено"
                    zeroed += 1
            if zeroed:
                self.note("typGRP.md §6.8",
                          f"«{label}»: условие «{key}» во входных данных не подтверждено — блок "
                          f"признан неприменимым, {zeroed} задач обнулены, структура сохранена. "
                          f"При подтверждении условия срок РС сдвигается вправо на длительность "
                          f"блока.")

    # ==================================================================
    # 5. Состав объектов
    # ==================================================================
    def object_blocks(self) -> dict[str, tuple[int, list[int]]]:
        """Объекты этапа: наименование заголовка → (индекс, индексы потомков)."""
        stage = self.one("СМР 1-й Этап")
        if stage is None:
            raise SystemExit("В шаблоне не найден блок «СМР 1-й Этап» — каркас изменился.")
        out = {}
        lvl = self.rows[stage]["lvl"] + 1
        for j in self.subtree(stage):
            if self.rows[j]["lvl"] == lvl:
                out[self.rows[j]["name"]] = (j, self.subtree(j))
        return out

    def check_stages(self) -> None:
        stages = self.p.get("этапы_ввода") or [{"номер": 1}]
        if len(stages) > 1:
            raise SystemExit(
                f"Этапов ввода в ТЭП: {len(stages)}. Шаблон v2 содержит один этап, а правила "
                f"разнесения объектов, нулевого цикла и благоустройства по этапам владельцем "
                f"не заданы (typGRP.md §14 п. 7). Расчёт остановлен — CLAUDE.md §0.3.")

    def configure_corpuses(self) -> None:
        """Корпуса проекта из двух прототипов шаблона — по составу остекления.

        Решение владельца 01.08.2026: прототипом для корпуса с витражами служит
        блок «Корпус 2», для прочих — «Корпус 1». Даты прототипов не наследуются:
        монолит пересобирается по этажности (§6), финиши выводятся от своего ЯКОРЬ.
        """
        blocks = self.object_blocks()
        proto_names = [n for n in blocks if n.startswith("Корпус")]
        if len(proto_names) < 2:
            raise SystemExit("В шаблоне ожидались два прототипа корпуса — каркас изменился.")
        protos = {}
        for n in proto_names:
            i, kids = blocks[n]
            pref = strip_prefix(self.rows[kids[0]]["name"])[0] if kids else ""
            protos[pref] = [dict(self.rows[j], links=list(self.rows[j]["links"]))
                            for j in [i] + kids]

        pvc_proto, stained_proto = protos.get("К1"), protos.get("К2")
        first, last = blocks[proto_names[0]][0], blocks[proto_names[-1]][1][-1]

        # Соответствие «ключ прототипа → ключи сгенерированных задач». Нужно для
        # внешних связей: агрегирующие вехи раздела 1 ссылаются на задачи корпусов.
        expand: dict[str, list[str]] = {}
        generated: list[dict] = []
        for n, corpus in enumerate(self.p["корпуса"], start=1):
            gl = corpus.get("остекление", {})
            proto = stained_proto if (gl.get("витражи") and not gl.get("пвх")) else pvc_proto
            src_pref = "К2" if proto is stained_proto else "К1"
            pref = f"К{n}"
            self.corpus_prefix[corpus["код"]] = pref
            keymap = {r["key"]: (r["key"] if r["key"].startswith("n") else self.newkey())
                      for r in proto}
            for r in proto:
                clone = dict(r)
                clone["key"] = keymap[r["key"]]
                clone["links"] = [(keymap.get(t, t), k, lg) for t, k, lg in r["links"]]
                if r is proto[0]:
                    clone["name"] = f"Корпус {n} ( {corpus['этажей_надземных']} этажей )"
                else:
                    clone["name"] = re.sub(rf"^{src_pref}\.", f"{pref}.", r["name"])
                clone["src"] = f"прототип {src_pref}, typGRP.md §12.1"
                generated.append(clone)
                expand.setdefault(r["key"], []).append(clone["key"])
            self.why(f"Корпус {corpus['код']}",
                     f"{corpus['этажей_надземных']} этажей, прототип {src_pref} "
                     f"({'витражи' if src_pref == 'К2' else 'ПВХ'})",
                     "typGRP.md §12.1", "средняя",
                     "Состав задач взят из прототипа; монолит пересобран по этажности, "
                     "финиши выведены от собственного ЯКОРЬ")

        self.rows[first:last + 1] = generated

        # Внешние связи на задачи прототипов размножаются по сгенерированным корпусам.
        proto_keys = set(expand)
        dropped = 0
        gen_keys = {r["key"] for r in generated}
        for r in self.rows:
            if r["key"] in gen_keys:
                continue
            new_links, changed = [], False
            for tgt, kind, lag in r["links"]:
                if tgt in proto_keys:
                    changed = True
                    for k in expand[tgt]:
                        new_links.append((k, kind, lag))
                    if not expand[tgt]:
                        dropped += 1
                else:
                    new_links.append((tgt, kind, lag))
            if changed:
                r["links"] = new_links
        if dropped:
            self.note("typGRP.md §12.1",
                      f"{dropped} связей шаблона вели на задачи прототипа, не попавшего в состав "
                      f"проекта, и отброшены вместе с ним.")
        if len(self.p["корпуса"]) > 2:
            self.note("typGRP.md §14 п. 7",
                      f"Проект содержит {len(self.p['корпуса'])} корпусов. Правила нумерации и "
                      f"состава вех при 3+ корпусах владельцем не подтверждены — блоки развёрнуты "
                      f"тиражированием прототипов по составу остекления.")

    def configure_zero_cycle(self) -> None:
        """Фундаменты и подземный монолит разбиты по корпусам — шаблон от 02.08.2026.

        До этого нулевой цикл был общим целиком: одна плита «Фундаменты Корпус» на
        всю очередь. Теперь у каждого корпуса своя плита и свой блок монолита ниже
        0.000, а `STD-ZC-005` применяется к этажности **своего** корпуса — этим
        закрыт `R-44`. Земляные работы, ограждение котлована и свайное поле
        остаются общими на очередь.
        """
        blocks = ("Фундаменты", "По договору Фундаменты Корпус"), \
                 ("Монолитные конструкции ниже отм 0,000",
                  "По договору Монолитные конструкции ниже отм 0,000")
        # old_key → (номер блока, порядковый номер корпуса в шаблоне)
        origin: dict[str, tuple[int, int]] = {}
        # (номер блока, номер корпуса проекта) → новый ключ
        made: dict[tuple[int, int], str] = {}
        gen_keys: set[str] = set()

        for bi, (head_name, bare) in enumerate(blocks):
            head = self.one(head_name)
            if head is None:
                continue
            corpus_rows = [j for j in self.subtree(head)
                           if strip_prefix(self.rows[j]["name"])[0].startswith("К")
                           and strip_prefix(self.rows[j]["name"])[1].startswith(bare)]
            if not corpus_rows:
                continue
            # Строки корпусов берутся из шаблона 1:1, как и блоки корпусов: у К1 и К2
            # разные предшественники (плита К2 ждёт свайное поле, плита К1 — нет),
            # и единый прототип эту разницу потерял бы.
            protos = [dict(self.rows[j], links=list(self.rows[j]["links"]))
                      for j in corpus_rows]
            for i, j in enumerate(corpus_rows):
                origin[self.rows[j]["key"]] = (bi, i)
            pos = corpus_rows[0]

            generated = []
            for n, _ in enumerate(self.p["корпуса"], start=1):
                src = protos[min(n - 1, len(protos) - 1)]
                clone = dict(src)
                clone["key"] = self.newkey()
                clone["links"] = list(src["links"])
                clone["name"] = re.sub(r"^К\d+\.", f"К{n}.", src["name"])
                clone["src"] = "покорпусный нулевой цикл, шаблон 02.08.2026"
                generated.append(clone)
                made[(bi, n)] = clone["key"]
            gen_keys.update(r["key"] for r in generated)

            for j in reversed(corpus_rows):
                del self.rows[j]
            self.rows[pos:pos] = generated

        if not origin:
            return

        # Внутри сгенерированных строк ссылка на покорпусную строку ведёт на СВОЙ
        # корпус (монолит ниже 0.000 ждёт плиту своего корпуса). Снаружи —
        # размножается на все корпуса: агрегирующие вехи собирают очередь целиком.
        n_corp = len(self.p["корпуса"])
        own: dict[str, int] = {}
        for (bi, n), key in made.items():
            own[key] = n
        for r in self.rows:
            new_links, changed = [], False
            for tgt, kind, lag in r["links"]:
                if tgt not in origin:
                    new_links.append((tgt, kind, lag))
                    continue
                changed = True
                bi, _ = origin[tgt]
                if r["key"] in own:
                    new_links.append((made[(bi, own[r["key"]])], kind, lag))
                else:
                    new_links.extend((made[(bi, n)], kind, lag)
                                     for n in range(1, n_corp + 1) if (bi, n) in made)
            if changed:
                r["links"] = new_links

        self.why("Нулевой цикл", f"фундаменты и подземный монолит — по {len(self.p['корпуса'])} "
                                 f"корпусам, земляные работы и сваи общие",
                 "typGRP.md §10.2", "высокая",
                 "Шаблон от 02.08.2026 разбил плиту и подземный монолит по корпусам: "
                 "STD-ZC-005 применяется к этажности своего корпуса (R-44 закрыт)")

    def configure_parking(self) -> None:
        """Паркинг несёт общий нулевой цикл, ИТП и точку КЛ_ПАРК (typGRP.md §10.2).

        При отсутствии паркинга в проекте эти блоки переносятся в первый корпус с
        его префиксом — bindings.md §1.1, RАСХОЖДЕНИЯ R-05/R-08. Прочие задачи
        паркинга (автостоянка, кладовые, МОП подземного уровня) снимаются.
        """
        blocks = self.object_blocks()
        park_name = next((n for n in blocks if n.startswith("Паркинг")), None)
        if park_name is None:
            return
        i, kids = blocks[park_name]
        self.parking_prefix = strip_prefix(self.rows[kids[0]]["name"])[0] if kids else "П1"

        if self.p.get("паркинг", {}).get("есть", False):
            return

        # Сохраняются блоки, которые в шаблоне лежат в паркинге, но относятся ко
        # всей очереди: общий нулевой цикл, кладка (точка КЛ_ПАРК), ИТП и отделка
        # техпомещений (от неё отсчитывается монтаж ИТП).
        keep = ("Нулевой цикл", "Общестроительные работы", "ЦТП, ИТП с УУТЭ",
                "Технические помещения")
        target = "К1"

        # Проход 1 — только отбор, БЕЗ изменения уровней: ancestors() читает
        # уровни, и правка их по ходу отвязала бы потомков от сохраняемого блока.
        roots = [j for j in kids if strip_prefix(self.rows[j]["name"])[1] in keep]
        keep_idx: list[int] = []
        for j in roots:
            keep_idx.append(j)
            keep_idx.extend(self.subtree(j))
        keep_set = sorted(set(keep_idx))
        dropped = len(kids) - len(keep_set)

        # Проход 2 — перенос. Каждый сохранённый корень встаёт на уровень 4
        # (ребёнок объекта), его потомки сдвигаются на ту же величину.
        shift = {j: 4 - self.rows[j]["lvl"] for j in roots}
        kept: list[dict] = []
        for j in keep_set:
            row = self.rows[j]
            root = max((r for r in roots if r <= j), default=None)
            row["lvl"] += shift.get(root, 0)
            row["name"] = f"{target}. {strip_prefix(row['name'])[1]}"
            kept.append(row)

        del self.rows[i:i + 1 + len(kids)]
        corpus1 = self.one("Надземная часть здания", prefix=target)
        pos = corpus1 if corpus1 is not None else 0
        self.rows[pos:pos] = kept
        self.parking_prefix = target
        self.note("R-05 / R-08",
                  f"Паркинг в проекте отсутствует. Шаблон v2 такого варианта не содержит: общий "
                  f"нулевой цикл, ИТП и точка отсчёта отделки КЛ_ПАРК живут в блоке паркинга. "
                  f"Сохранённые блоки перенесены в корпус К1 с его префиксом, {dropped} задач "
                  f"паркинга (автостоянка, кладовые, МОП подземного уровня) сняты. Точка КЛ_ПАРК "
                  f"заменена на ЯКОРЬ_ОГР корпуса. Требует правил владельца.")

    # ==================================================================
    # 6. Нормативы и поэтажная развёртка
    # ==================================================================
    def set_dur(self, i: int | None, value: int, source: str) -> None:
        if i is not None:
            self.rows[i]["dur"] = value
            self.rows[i]["src"] = source

    @staticmethod
    def _live_preds(key: str, dead: dict[str, list[tuple[str, str, int]]],
                    seen: set[str] | None = None) -> list[tuple[str, str, int]]:
        """Предшественники удаляемой строки, разрешённые до живых строк.

        Если предшественник удаляемой строки сам удаляется, спуск продолжается
        вглубь: иначе цепочка «живой → удаляемая → удаляемая → потребитель»
        оставила бы потребителя вовсе без связей.
        """
        seen = seen if seen is not None else set()
        out: list[tuple[str, str, int]] = []
        for pt, pk, plg in dead.get(key, []):
            if pt in seen:
                continue
            seen.add(pt)
            if pt in dead:
                out.extend(Build._live_preds(pt, dead, seen))
            else:
                out.append((pt, pk, plg))
        return out

    def drop_rows(self, roots: list[int], reason: str) -> int:
        """Удаляет строки `roots` вместе с потомками и переписывает внешние связи.

        CLAUDE.md §12 запрещает молчаливое отбрасывание связей. finalize()
        отбросил бы ссылки на удалённые строки без следа, поэтому решение
        принимается здесь и записывается в «Обоснование»:
          · у потребителя остаются другие предшественники → битая связь снимается;
          · других предшественников нет → потребитель наследует предшественников
            удаляемой строки (транзитивно, до первой живой).
        """
        doomed: set[int] = set()
        for i in roots:
            doomed.add(i)
            doomed.update(self.subtree(i))
        if not doomed:
            return 0

        dead = {self.rows[i]["key"]: list(self.rows[i]["links"]) for i in doomed}
        cut: list[str] = []
        rewired: list[str] = []
        for i, r in enumerate(self.rows):
            if i in doomed:
                continue
            lost = [ln for ln in r["links"] if ln[0] in dead]
            if not lost:
                continue
            kept = [ln for ln in r["links"] if ln[0] not in dead]
            if kept:
                r["links"] = kept
                cut.append(r["name"])
            else:
                inherited: list[tuple[str, str, int]] = []
                for t, _, _ in lost:
                    for ln in self._live_preds(t, dead):
                        if ln not in inherited:
                            inherited.append(ln)
                r["links"] = inherited
                rewired.append(r["name"])

        for i in sorted(doomed, reverse=True):
            del self.rows[i]

        self.why(f"Удаление блока ({reason})",
                 f"{len(doomed)} строк снято",
                 f"{reason} · CLAUDE.md §12", "средняя",
                 f"Связей снято у сохранённых задач: {len(cut)}"
                 + (f" — {'; '.join(cut[:6])}" if cut else "")
                 + f". Связей переопределено на предшественников удалённых: {len(rewired)}"
                 + (f" — {'; '.join(rewired[:6])}" if rewired else "")
                 + ". Ни одна зависимость не отброшена молча.")
        return len(doomed)

    def apply_standards(self) -> None:
        zc = self.p["нулевой_цикл"]
        wall = zc["ограждение_котлована"].lower()
        pk = self.parking_prefix

        # --- земляные работы (STD-ZC-002) ---
        if wall not in Std.EARTH:
            self.note("standards.md §6", f"Тип ограждения котлована «{wall}» не распознан — "
                                         f"принят трубошпунт по умолчанию.")
            wall = "трубошпунт"
        earth_days = Std.EARTH[wall][0]
        self.set_dur(self.one("По договору Земляные работы", pk), earth_days, Std.EARTH_SRC)
        self.why("Земляные работы", f"{earth_days} дн", Std.EARTH_SRC, "высокая",
                 f"По типу ограждения «{wall}», не по этажности подземной части (DEC-03)")

        # --- ограждение котлована (STD-ZC-001) ---
        if wall == "отвал":
            self.note("standards.md §5 — открытый вопрос",
                      "Вход «отвал» (без ограждения котлована): STD-ZC-001 норматива для этого "
                      "варианта не содержит. Задача сохранена со связями трубошпунта — переносить "
                      "её без правила привязки нельзя (CLAUDE.md §0.3).")
            wall_days, _, wall_lag = Std.PIT_WALL["трубошпунт"]
        else:
            wall_days, _, wall_lag = Std.PIT_WALL[wall]
        i = self.one("По договору Ограждения котлована", pk)
        self.set_dur(i, wall_days, Std.PIT_WALL_SRC)
        earth = self.one("По договору Земляные работы", pk)
        if i is not None and earth is not None:
            self.rows[i]["links"] = [(self.rows[earth]["key"], "НН", wall_lag)]
        self.why("Ограждение котлована", f"{wall_days} дн, НН {wall_lag:+d} дн от земляных",
                 Std.PIT_WALL_SRC, "средняя" if wall == "отвал" else "высокая", f"Тип: {wall}")

        # --- свайное основание (STD-ZC-003/004) ---
        pile_days, trace = pile_duration(zc.get("сваи", []))
        for nm in ("По договору Свайное основание БНС (при наличии)",
                   "По договору Свайное основание Забивные (при наличии)"):
            self.set_dur(self.one(nm, pk), pile_days, "standards.md §7 STD-ZC-003/004")
        self.why("Свайное основание", f"{pile_days} дн", "standards.md §7 STD-ZC-003/004",
                 "средняя", " · ".join(trace) + " · статус ⚠, калибровки не проходило")
        if pile_days > Std.PILE_WARN_DAYS[0]:
            msg = (f"Свайное основание {pile_days} дн превышает порог {Std.PILE_WARN_DAYS[0]} дн. "
                   f"Рекомендуется увеличить число буровых установок.")
            self.notices.append(msg)
            self.note("DEC-14", msg)

        # --- фундаменты и монолит ниже 0.000 (общие на очередь, typGRP.md §10.2) ---
        # --- фундаментная плита: по этажности СВОЕГО корпуса (R-44 закрыт) ---
        rafts = []
        for n, corpus in enumerate(self.p["корпуса"], start=1):
            nf = corpus["этажей_надземных"]
            raft, src = (Std.RAFT_TALL if nf > 45 else Std.RAFT)
            self.set_dur(self.one("По договору Фундаменты Корпус", f"К{n}"), raft, src)
            rafts.append(f"{corpus['код']} ({nf} эт.) — {raft} дн")
        self.why("Фундаментные плиты корпусов", " · ".join(rafts),
                 "standards.md §8 STD-ZC-005", "высокая",
                 "Шаблон от 02.08.2026 разбил плиту по корпусам, поэтому порог 45 этажей "
                 "применяется к каждому корпусу отдельно — прежняя неоднозначность R-44 снята")

        lv = zc.get("этажей_подземных", 1)
        below = Std.BELOW_ZERO[0] * lv
        for i in self.find("По договору Монолитные конструкции ниже отм 0,000",
                           None, exact=False):
            self.set_dur(i, below, Std.BELOW_ZERO[1])
        self.why("Монолит ниже 0.000", f"{below} дн на объект", Std.BELOW_ZERO[1], "высокая",
                 f"{lv} подземн. этаж(а) × {Std.BELOW_ZERO[0]} дн; развёрнут по корпусам "
                 f"и паркингу отдельными задачами")

        park = self.p.get("паркинг", {})
        if park.get("есть"):
            plv = park.get("этажей_подземных", 1)
            self.set_dur(self.one("По договору Фундаменты Паркинг", pk),
                         Std.PARKING_RAFT[0] + Std.PARKING_LEVEL[0] * plv, Std.PARKING_RAFT[1])
            self.why("Фундаменты паркинга",
                     f"{Std.PARKING_RAFT[0] + Std.PARKING_LEVEL[0] * plv} дн",
                     "standards.md §10 STD-MON-002", "высокая",
                     f"ФП 45 дн + {plv} подземн. этаж(а) × 45 дн")
            self.set_dur(self.one("По договору Кладка наружных и внутренних стен и перегородок",
                                  pk), Std.PARKING_MASONRY[0], Std.PARKING_MASONRY[1])
            self.set_dur(self.one("Устройство гидроизоляции", pk),
                         Std.PARKING_ROOF[0], Std.PARKING_ROOF[1])

        # --- отделка паркинга (STD-OTD-003/005/006/007) ---
        for nm, (val, src) in (("По договору Технические помещения", Std.TECH_ROOMS),
                               ("По договору Автостоянка (в т.ч. Рампа)", Std.PARKING_LOT),
                               ("По договору Кладовые", Std.STORAGE),
                               ("По договору МОП подземного уровня", Std.UNDERGROUND_MOP)):
            self.set_dur(self.one(nm, pk), val, src)
        self.why("Отделка паркинга",
                 f"техпомещения {Std.TECH_ROOMS[0]} · автостоянка {Std.PARKING_LOT[0]} · "
                 f"кладовые {Std.STORAGE[0]} · МОП {Std.UNDERGROUND_MOP[0]} дн",
                 "standards.md §15", "средняя",
                 "Техпомещения 90 дн отменяют DEC-08 (было 60) — R-06")

        # --- ИТП (STD-VIS-001, BND-VIS-006) ---
        itp = self.one("По договору ЦТП, ИТП с УУТЭ", pk)
        tech = self.one("По договору Технические помещения", pk)
        if itp is not None:
            self.set_dur(itp, Std.ITP[0], Std.ITP[1])
            if tech is not None:
                self.rows[itp]["links"] = [(self.rows[tech]["key"], "НН", Bnd.LAG_ITP[0])]
                self.rows[itp]["comment"] = (f"BND-VIS-006: НН +{Bnd.LAG_ITP[0]} дн от старта "
                                             f"отделки техпомещений · {Std.ITP[1]}")
        self.why("ЦТП/ИТП", f"{Std.ITP[0]} дн, НН +{Bnd.LAG_ITP[0]} дн от техпомещений",
                 Std.ITP[1], "средняя",
                 "R-03/R-04: отменяют DEC-06 — было 200 дн и НН +90 от кладки перегородок")

        # --- отделка техпомещений: НН +21 от старта кладки паркинга (BND-VIS-004) ---
        masonry_pk = self.one("По договору Кладка наружных и внутренних стен и перегородок", pk)
        if tech is not None and masonry_pk is not None:
            self.rows[tech]["links"] = [(self.rows[masonry_pk]["key"], "НН",
                                         Bnd.LAG_TECH_ROOMS[0])]
            self.rows[tech]["comment"] = (f"BND-VIS-004: НН +{Bnd.LAG_TECH_ROOMS[0]} дн от старта "
                                          f"кладки паркинга · {Std.TECH_ROOMS[1]}")

        # --- ВИС паркинга (STD-VIS-007) ---
        for i in self.find("По договору", pk, exact=False):
            nm = strip_prefix(self.rows[i]["name"])[1]
            if any(nm.startswith(x) for x in ("По договору Отопление", "По договору Вентиляция",
                                              "По договору Кондиционирование",
                                              "По договору Электроснабжение")):
                self.set_dur(i, Std.VIS_PARKING[0], Std.VIS_PARKING[1])

        # --- фиксированные нормативы корпусов ---
        for pref in self.corpus_prefix.values():
            self.set_dur(self.one("По договору МОКАП фасада", pref), *Std.FACADE_MOCKUP)
            self.set_dur(self.one("По договору Вертикальный транспорт (лифты,подъемное "
                                  "оборудование)", pref), *Std.ELEVATORS)
            self.set_dur(self.one("Закрыт ВРЕМЕННЫЙ тепловой контур по корпусу "
                                  "(при необходимости)", pref), *Std.TEMP_CONTOUR)
        self.why("МОКАП фасада", f"{Std.FACADE_MOCKUP[0]} дн", Std.FACADE_MOCKUP[1], "высокая",
                 "DEC-12: включается по умолчанию, по каждому корпусу")
        self.why("Лифты", f"{Std.ELEVATORS[0]} дн", Std.ELEVATORS[1], "высокая", "DEC-11")

        # --- благоустройство (STD-BIO-*) ---
        bio = self.one("Благоустройство территории")
        if bio is not None:
            names = {"По договору Земляные работы": "STD-BIO-001",
                     "По договору Устройство твердых покрытий": "STD-BIO-002",
                     "По договору Озеленение": "STD-BIO-003",
                     "По договору Установка ограждения": "STD-BIO-004",
                     "По договору Устройство наружного освещения": "STD-BIO-005",
                     "По договору Монтаж МАФ": "STD-BIO-006",
                     "По договору Организация дорожного движения": "STD-BIO-007"}
            for j in self.subtree(bio):
                code = names.get(strip_prefix(self.rows[j]["name"])[1])
                if code:
                    self.rows[j]["dur"] = Std.BIO[code][0]
                    self.rows[j]["src"] = f"standards.md §17 {code}"
            self.why("Благоустройство", f"{Std.BIO_BLOCK[0]} дн на блок", Std.BIO_BLOCK[1],
                     "средняя",
                     "R-07: резервной строки нет — цепочка «земляные 60 → твёрдые покрытия 120» "
                     "даёт ровно 180 дн, DEC-13 отменён")

    def rebuild_monolith(self, corpus: dict) -> None:
        pref = self.corpus_prefix[corpus["код"]]
        head = self.one("Монтаж монолитных конструкций выше отм. 0.000", pref)
        if head is None:
            return
        n = corpus["этажей_надземных"]
        first, typical, roof = monolith_floor_durations(n, corpus.get("сложный_конструктив", False))
        k = typical[0] if typical else Std.MON_TYPICAL[0]

        kids = self.subtree(head)
        old_keys = [self.rows[j]["key"] for j in kids]
        head_links = list(self.rows[kids[0]]["links"]) if kids else []
        lvl = self.rows[head]["lvl"] + 1
        del self.rows[kids[0]:kids[-1] + 1]

        new, prev = [], None
        for i in range(1, n + 1):
            key = self.newkey()
            new.append({"key": key, "lvl": lvl, "name": f"{pref}. {i} этаж Монолит",
                        "dur": first if i == 1 else k,
                        "links": [(prev, "ОН", 0)] if prev else list(head_links),
                        "comment": f"STD-MON-001: {'первый этаж' if i == 1 else f'K = {k} дн'}",
                        "tpl_start": "", "src": "standards.md §9 STD-MON-001"})
            prev = key
        roof_key = self.newkey()
        new.append({"key": roof_key, "lvl": lvl, "name": f"{pref}. Кровля/Парапет Монолит",
                    "dur": roof, "links": [(prev, "ОН", 0)],
                    "comment": "STD-MON-001: кровля/парапет. ЯКОРЬ — последний куб бетона",
                    "tpl_start": "", "src": "standards.md §9 STD-MON-001"})
        self.rows[kids[0]:kids[0]] = new

        # Лаги пересчитываются, а не копируются — typGRP.md §12.2.
        # Прототип содержал столько же этажей, сколько его исходный корпус, поэтому
        # ссылки на «6 этаж» и «15 этаж» переносятся на min(6, N) и min(15, N).
        floor6 = new[min(6, n) - 1]["key"]
        floor15 = new[min(15, n) - 1]["key"]
        proto_n = len(old_keys) - 1
        old6 = old_keys[min(6, proto_n) - 1] if proto_n >= 1 else None
        old15 = old_keys[min(15, proto_n) - 1] if proto_n >= 15 else None
        remap = {}
        for idx, ok in enumerate(old_keys):
            if ok == old6:
                remap[ok] = floor6
            elif ok == old15:
                remap[ok] = floor15
            elif idx == len(old_keys) - 1:
                remap[ok] = roof_key
            else:
                remap[ok] = roof_key      # ссылка на исчезнувший этаж → ЯКОРЬ
        changed = 0
        for r in self.rows:
            fixed, hit = [], False
            for tgt, kind, lag in r["links"]:
                if tgt in remap:
                    tgt, hit = remap[tgt], True
                fixed.append((tgt, kind, lag))
            if hit:
                r["links"] = fixed
                changed += 1

        total = first + k * (n - 1) + roof
        self.why(f"Монолит корпуса {corpus['код']}",
                 f"{total} дн ({first} + {k} × {n - 1} + {roof})",
                 "standards.md §9 STD-MON-001", "высокая",
                 f"K = {k} дн ({'сложный конструктив или N > 60' if k == 7 else 'типовой'}); "
                 f"{changed} связей перепривязано на этажи min(6, {n}) = {min(6, n)} и "
                 f"min(15, {n}) = {min(15, n)} — typGRP.md §12.2")

    # ==================================================================
    # 7. Сетевой расчёт
    # ==================================================================
    def build_nodes(self) -> dict[str, Node]:
        """Сеть включает и суммарные строки — они участвуют свёрточным узлом.

        Связи шаблона ссылаются на суммарные строки (80 штук), и ссылка обязана
        сохраниться ровно такой (DEC-26). Поэтому суммарная строка присутствует в
        расчёте: своей длительности и связей не имеет, старт = минимум стартов
        потомков, финиш = максимум финишей (typGRP.md §2.2 п. 4).
        """
        nodes: dict[str, Node] = {}
        for i, r in enumerate(self.rows):
            if self.is_summary(i):
                kids = [self.rows[j]["key"] for j in self.subtree(i)
                        if self.rows[j]["lvl"] == r["lvl"] + 1]
                nodes[r["key"]] = Node(key=r["key"], duration=0, rollup=kids)
                continue
            n = Node(key=r["key"], duration=r["dur"] or 0, links=list(r["links"]))
            if not n.links and r.get("tpl_start"):
                n.anchor_start = dparse(r["tpl_start"]) + timedelta(days=self.shift)
            nodes[r["key"]] = n
        return nodes

    def schedule(self) -> dict[str, Node]:
        nodes = self.build_nodes()
        forward_pass(nodes, self.start)
        backward_pass(nodes)
        return nodes

    def report_anchors(self, nodes: dict[str, Node]) -> None:
        anchored = [k for k, n in nodes.items() if n.anchor_start and not n.links]
        if not anchored:
            return
        self.note("CLAUDE.md §9",
                  f"{len(anchored)} задач-листьев не имеют предшественников ни у себя, ни у "
                  f"предков, и удерживаются календарной датой шаблона (регламентные вехи фазы A, "
                  f"контрольные вехи проектирования). Это помеченные ограничения; при поступлении "
                  f"правил привязки они заменяются связями.")
        self.why("Календарные якоря", f"{len(anchored)} задач",
                 "typGRP.md §6.0, регламентные привязки и длительности", "средняя",
                 "Пересчёту по standards.md не подлежат — длительности берутся как есть")

    # ==================================================================
    # 8. Тепловой контур и отделка — bindings.md §3.6, §3.8
    # ==================================================================
    def thermal(self, nodes: dict[str, Node]) -> None:
        facade = self.p["фасад"]["тип"]
        pk = self.parking_prefix
        has_parking = self.p.get("паркинг", {}).get("есть", False)

        # ИТП общий, лежит в паркинге (BND-VIS-006)
        itp = self.one("По договору ЦТП, ИТП с УУТЭ", pk)
        itp_finish = nodes[self.rows[itp]["key"]].finish if itp is not None else None

        for corpus in self.p["корпуса"]:
            pref = self.corpus_prefix[corpus["код"]]
            gl = corpus.get("остекление", {})
            n_fl = corpus["этажей_надземных"]

            roof = self.one("Кровля/Парапет Монолит", pref)
            if roof is None:
                continue
            roof_key = self.rows[roof]["key"]
            anchor = nodes[roof_key].finish

            ext = self.one("По договору Кладка наружных стен", pref)
            spk = [i for i in self.find("По договору Монтаж светопрозрачных конструкций",
                                        pref, exact=False)]
            pvc = next((i for i in spk if "ПВХ" in self.rows[i]["name"]), None)
            stained = next((i for i in spk if "Витраж" in self.rows[i]["name"]), None)

            # ЯКОРЬ_ОГР — BND-KLD-001 / BND-SPK-003
            if gl.get("витражи_на_всю_высоту") or ext is None:
                src = stained if stained is not None else roof
                anchor_ogr = nodes[self.rows[src]["key"]].start
                self.note("BND-SPK-003", f"{corpus['код']}: витражи на всю высоту — наружной "
                                         f"кладки нет, ЯКОРЬ_ОГР = старт монтажа витражей.",
                          corpus["код"])
            else:
                anchor_ogr = nodes[self.rows[ext]["key"]].start

            # Финиши остекления и фасада — от собственного ЯКОРЬ корпуса
            no_masonry = gl.get("витражи_на_всю_высоту") or ext is None
            pvc_off = Bnd.FIN_PVC_NO_MASONRY[0] if no_masonry else Bnd.FIN_PVC[0]
            pvc_fin = anchor + timedelta(days=pvc_off) if pvc is not None else None
            st_fin = anchor + timedelta(days=Bnd.FIN_STAINED[0]) if stained is not None else None
            fac_days = (Bnd.FIN_FACADE_MODULAR[0]
                        if facade.lower() in ("модульный", "панельный") else Bnd.FIN_FACADE[0])
            fac_fin = anchor + timedelta(days=fac_days)

            # Отопление — контур для пуска тепла: 250 дн от старта блока ВИС (R-01)
            loop = self.one("По договору Отопление (контур для пуска тепла)", pref)
            loop_finish = None
            if loop is not None:
                self.rows[loop]["dur"] = Std.HEAT_LOOP[0]
                self.rows[loop]["comment"] = (f"{Std.HEAT_LOOP[1]} · ОН +{Bnd.LAG_HEAT[0]} дн "
                                              f"к вехе «Пуск тепла» (R-02)")
                loop_finish = (nodes[self.rows[loop]["key"]].start
                               + timedelta(days=Std.HEAT_LOOP[0]))

            # Старт отделки — НН +90 дн от старта кладки перегородок СВОЕГО корпуса
            # (шаблон от 02.08.2026). Прежняя привязка к финишу кладки паркинга
            # (КЛ_ПАРК, ОН +90) отменена вместе с DEC-28.
            part = self.one("По договору Кладка перегородок", pref)
            if part is not None:
                base = nodes[self.rows[part]["key"]].start
                base_label = "старт кладки перегородок корпуса"
            else:
                base = anchor_ogr
                base_label = "ЯКОРЬ_ОГР"
                self.note("BND-OTD-002", f"{corpus['код']}: кладки перегородок в блоке нет — "
                                         f"старт отделки отсчитан от ЯКОРЬ_ОГР.", corpus["код"])

            res = compute_thermal_and_fit(
                anchor=anchor, anchor_ogr=anchor_ogr, facade_type=facade,
                pvc_finish=pvc_fin, stained_finish=st_fin, facade_finish=fac_fin,
                fit_base=base, fit_base_label=base_label,
                itp_finish=itp_finish, heat_loop_finish=loop_finish,
                masonry_finish=(nodes[self.rows[ext]["key"]].finish
                                if ext is not None else None),
            )
            self.assumptions.extend(
                Assumption(a.code, a.text, corpus["код"]) for a in res.assumptions)

            def link_to_anchor(i: int | None, start: date, dur: int, comment: str) -> None:
                """Привязка связью ОН к вехе монолита кровли (ЯКОРЬ) с численным лагом.

                Даты напрямую не проставляются: чек-лист CLAUDE.md §9 требует
                привязки к вехам монолита, а не к датам.
                """
                if i is None:
                    return
                lag = (start - anchor).days
                self.rows[i]["dur"] = dur
                self.rows[i]["links"] = [(roof_key, "ОН", lag)]
                self.rows[i]["tpl_start"] = ""
                self.rows[i]["comment"] = f"{comment} · ОН от ЯКОРЬ, лаг {lag:+d} дн"

            link_to_anchor(self.one("Закрыт тепловой контур по корпусу", pref), res.tc_perm, 0,
                           f"TC_perm, BND-TC-001, тип фасада {facade}")
            vtk = self.one("Закрыт ВРЕМЕННЫЙ тепловой контур по корпусу (при необходимости)", pref)
            if res.temp_contour:
                link_to_anchor(vtk, res.temp_contour[0], Std.TEMP_CONTOUR[0],
                               "ВТК, BND-TC-002: финиш 30.09 года старта отделки, 45 дн")
            elif vtk is not None:
                del self.rows[vtk]
                self.note("BND-TC-002", "ВТК не развёрнут: условие не выполнено либо "
                                        "развёртывание не даёт эффекта.", corpus["код"])

            heat_i = self.one("Пуск тепла корпус", pref)
            link_to_anchor(heat_i, res.heat, 0,
                           f"Пуск тепла, BND-TC-003: max(TC_eff, ИТП, отопление) "
                           f"+ {Bnd.LAG_HEAT[0]} дн (R-02)")

            # --- отделка: пишется в задачи-листья, не в суммарные строки ---
            fit_leaves = []
            for nm in ("По договору Вестибюль", "По договору Лифтовые и квартирные холлы",
                       "По договору Л/клетки и тамбур-шлюзы",
                       "По договору Коммерческие помещения",
                       "По договору Подготовка под чистовую отделку квартир",
                       "По договору Чистовая отделка квартир"):
                i = self.one(nm, pref)
                if i is not None:
                    fit_leaves.append(i)
            for i in fit_leaves:
                link_to_anchor(i, res.fit_start, res.fit_duration,
                               f"BND-OTD-002/003: старт от {base_label} +{Bnd.LAG_FIT[0]} через "
                               f"сезонный гейт, длительность {res.fit_duration} дн (минимум 210)")
                self.fit_tasks.append((self.rows[i]["key"], corpus["код"]))

            # --- раздел ТХ: ОН от вехи пуска тепла (BND-OTD-004) ---
            tx = self.one("По договору Отделочные работы, раздел ТХ", pref, exact=False)
            if tx is not None and heat_i is not None:
                dur = (res.fit_finish - res.heat).days
                if dur > 0:
                    self.rows[tx]["dur"] = dur
                    self.rows[tx]["links"] = [(self.rows[heat_i]["key"], "ОН", 0)]
                    self.rows[tx]["tpl_start"] = ""
                    self.rows[tx]["comment"] = ("BND-OTD-004: ОН от вехи «Пуск тепла корпус», "
                                                "финиш совпадает с финишем отделки корпуса")

            # --- финиши, привязанные к ЯКОРЬ; длительность из окна ---
            def finish_at_anchor(i: int | None, offset: int, label: str, src: str) -> None:
                if i is None:
                    return
                node = nodes.get(self.rows[i]["key"])
                if node is None or node.start is None:
                    return
                fin = anchor + timedelta(days=offset)
                dur = (fin - node.start).days
                if dur <= 0:
                    self.note(src, f"{corpus['код']}: расчётный финиш «{label}» ({dfmt(fin)}) "
                                   f"не позже старта ({dfmt(node.start)}). Длительность шаблона "
                                   f"сохранена.", corpus["код"])
                    return
                self.rows[i]["dur"] = dur
                self.rows[i]["comment"] = f"{label}: финиш ЯКОРЬ +{offset} дн — {src}"

            finish_at_anchor(ext, Bnd.FIN_MASONRY_EXT[0], "Кладка наружных стен",
                             Bnd.FIN_MASONRY_EXT[1])
            finish_at_anchor(self.one("По договору Кладка перегородок", pref),
                             Bnd.FIN_MASONRY_INT[0], "Кладка перегородок", Bnd.FIN_MASONRY_INT[1])
            finish_at_anchor(pvc, pvc_off, "СПК ПВХ", Bnd.FIN_PVC[1])
            finish_at_anchor(stained, Bnd.FIN_STAINED[0], "Витражи", Bnd.FIN_STAINED[1])
            fac = self.one("По договору Монтаж фасадов", pref)
            finish_at_anchor(fac, fac_days, "Монтаж фасадов",
                             Bnd.FIN_FACADE[1] if fac_days == Bnd.FIN_FACADE[0]
                             else Bnd.FIN_FACADE_MODULAR[1])

            # --- ВИС: длительность выводится из финиша ЯКОРЬ +330 (DEC-23) ---
            vis_heads = ["Внутренние сантехнические системы", "Отопление",
                         "Вентиляция, дымоудаление", "Кондиционирование",
                         "Внутренние слаботочные системы", "Электроснабжение и электроосвещение"]
            changed = 0
            for hname in vis_heads:
                h = self.one(hname, pref)
                if h is None:
                    continue
                for j in self.subtree(h):
                    if self.is_summary(j) or j == loop:
                        continue
                    before = self.rows[j]["dur"]
                    finish_at_anchor(j, Bnd.FIN_VIS[0], f"ВИС {hname}", Bnd.FIN_VIS[1])
                    if self.rows[j]["dur"] != before:
                        changed += 1
            if changed:
                self.why(f"ВИС {corpus['код']}", f"{changed} задач пересчитано от ЯКОРЬ",
                         "bindings.md §3.7 BND-VIS-001/002 · DEC-23", "высокая",
                         f"Финиш ВИС = ЯКОРЬ +{Bnd.FIN_VIS[0]} дн "
                         f"({dfmt(anchor + timedelta(days=Bnd.FIN_VIS[0]))}). Статические "
                         f"длительности шаблона не переносятся — длительность выводится из окна "
                         f"«старт по вехе min(15, N) этаж → финиш по ЯКОРЬ»")

            # --- кровля: ОН от монолита кровли + SEA-02 ---
            for nm in ("Устройство кровли неэксплуатируемой", "Устройство кровли эксплуатируемой"):
                h = self.one(nm, pref)
                if h is None:
                    continue
                leaves = [j for j in self.subtree(h) if not self.is_summary(j)] or [h]
                for j in leaves:
                    self.rows[j]["links"] = [(roof_key, "ОН", 0)]
                    self.rows[j]["tpl_start"] = ""
                    base_d = self.rows[j]["dur"] or Std.ROOF_PLAIN[0]
                    adj = seasonal_duration(anchor, base_d, "SEA-02", 2.0)
                    self.rows[j]["dur"] = adj
                    self.rows[j]["comment"] = (
                        "BND-FAS-004: ОН от монолита кровли/парапета"
                        + (f" · SEA-02: {base_d} → {adj} дн, ×2,0 к попавшей части (DEC-20)"
                           if adj != base_d else ""))

            self.why(f"Тепловой контур {corpus['код']}",
                     f"TC_perm {dfmt(res.tc_perm)} · пуск тепла {dfmt(res.heat)}",
                     "bindings.md §3.6 BND-TC-001…003", "средняя", " · ".join(res.trace))
            self.why(f"Отделка {corpus['код']}",
                     f"{dfmt(res.fit_start)} → {dfmt(res.fit_finish)}, {res.fit_duration} дн",
                     "bindings.md §3.8 BND-OTD-002/003 · standards.md §15.1", "средняя",
                     f"ЯКОРЬ = {dfmt(res.anchor)} · ЯКОРЬ_ОГР = {dfmt(res.anchor_ogr)} · "
                     f"точка отсчёта — {base_label}")
            if res.overrun_days:
                self.notices.append(
                    f"{corpus['код']}: финиш отделки выходит за ЯКОРЬ + 365 на "
                    f"{res.overrun_days} дн. РВЭ и передача сдвинуты на ту же величину (DEC-09).")

    # ==================================================================
    # 9. Вход блока ЗОС — BND-ZOS-001, ОН −160 от финиша отделки
    # ==================================================================
    def wire_zos(self) -> None:
        i = self.one("Готовность к обмерам БТИ")
        if i is None or not self.fit_tasks:
            return
        self.rows[i]["links"] = [(k, "ОН", -Bnd.LAG_BTI[0]) for k, _ in self.fit_tasks]
        self.rows[i]["tpl_start"] = ""
        self.rows[i]["comment"] = (f"BND-ZOS-001 (DEC-29): ОН −{Bnd.LAG_BTI[0]} дн от финиша "
                                   f"отделки каждого корпуса этапа. Веха РВЭ определяется самым "
                                   f"поздним корпусом — CLAUDE.md §4; сдвиг отделки доходит до "
                                   f"РВЭ и передачи — DEC-09")
        self.why("Блок ЗОС → РВЭ", f"ОН −{Bnd.LAG_BTI[0]} дн от отделки "
                                   f"{len(self.fit_tasks)} задач",
                 "bindings.md §3.10 BND-ZOS-001 · DEC-29", "средняя",
                 "В шаблоне веха удерживалась связью на суммарные строки отделки; "
                 "правило сохранено, цели заменены на задачи-листья")

    # ==================================================================
    # 10. Финализация
    # ==================================================================
    def finalize(self, nodes: dict[str, Node]) -> list[dict]:
        summaries = self.summaries()
        key2id = {r["key"]: i + 1 for i, r in enumerate(self.rows)}
        out: list[dict] = []
        succ: dict[int, list[int]] = {}

        for i, r in enumerate(self.rows, start=1):
            key = r["key"]
            is_sum = key in summaries
            n = nodes.get(key)
            links = [(key2id[t], k, lg) for t, k, lg in r["links"] if t in key2id]
            # DEC-27: «Последователи» — только Ид., без кода связи и лага.
            if not is_sum:
                for pid, _, _ in links:
                    succ.setdefault(pid, []).append(i)
            out.append({
                "Вид работ": "",
                "Код классификатора": "",
                "Уровень структуры": r["lvl"],
                "Ид.": i,
                "Название задачи": r["name"],
                "% завершения": "" if is_sum else 0,
                "Длительность": "" if is_sum else f"{(r['dur'] or 0)} дней",
                "Начало": dfmt(n.start) if n and n.start else "",
                "Окончание": dfmt(n.finish) if n and n.finish else "",
                "Предшественники": "" if is_sum else fmt_links(links),
                "Последователи": "",
                "комментарий": ("КРИТИЧЕСКИЙ ПУТЬ · " if (n and n.critical and not is_sum) else "")
                               + r.get("comment", ""),
                "_critical": bool(n and n.critical) and not is_sum,
                "_source": r.get("src", ""),
            })
        for row in out:
            row["Последователи"] = "; ".join(str(x) for x in succ.get(row["Ид."], []))
        return out


# ======================================================================
# Запись Excel
# ======================================================================
def write_excel(path: Path, rows: list[dict], build: Build) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ГРП"

    head = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2F5597")
    sum_font = Font(bold=True)
    crit_fill = PatternFill("solid", fgColor="FFE0E0")

    ws.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        ws.cell(1, c).font = head
        ws.cell(1, c).fill = head_fill
        ws.cell(1, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    name_col = COLUMNS.index("Название задачи") + 1
    for r in rows:
        ws.append([r[c] for c in COLUMNS])
        i = ws.max_row
        if not r["Длительность"]:
            for c in range(1, len(COLUMNS) + 1):
                ws.cell(i, c).font = sum_font
        elif r["_critical"]:
            for c in range(1, len(COLUMNS) + 1):
                ws.cell(i, c).fill = crit_fill
        ws.cell(i, name_col).alignment = Alignment(indent=max(0, r["Уровень структуры"] - 1))

    for i, w in enumerate([12, 16, 9, 7, 62, 11, 13, 12, 12, 30, 24, 60], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    def sheet(title: str, header: list[str], data: list[list], widths: list[int]):
        s = wb.create_sheet(title)
        s.append(header)
        for c in range(1, len(header) + 1):
            s.cell(1, c).font = head
            s.cell(1, c).fill = head_fill
        for row in data:
            s.append(row)
        for i, w in enumerate(widths, start=1):
            s.column_dimensions[get_column_letter(i)].width = w
        s.freeze_panes = "A2"
        return s

    sheet("Обоснование", ["Блок", "Значение", "Источник", "Уверенность", "Комментарий"],
          [[x["Блок"], x["Значение"], x["Источник"], x["Уверенность"], x["Комментарий"]]
           for x in build.rationale], [34, 40, 46, 14, 110])

    assumptions = [[a.code, a.scope, a.text] for a in build.assumptions]
    assumptions.append(["DEC-16", "", "Правила сценариев не заданы — блок сценариев и "
                                      "вероятностный расчёт не рассчитаны (CLAUDE.md §8)."])
    sheet("Допущения", ["Код", "Область", "Формулировка"], assumptions, [24, 12, 130])

    crit = sum(1 for r in rows if r["_critical"])
    leaves = [r for r in rows if r["Длительность"]]
    sums = [r for r in rows if not r["Длительность"]]
    checks = [
        ("Структура", "Выведены все 12 колонок", True),
        ("Структура", "Уровень первой строки = 1", rows[0]["Уровень структуры"] == 1),
        ("Структура", "Уровень не растёт больше чем на 1",
         all(rows[i]["Уровень структуры"] - rows[i - 1]["Уровень структуры"] <= 1
             for i in range(1, len(rows)))),
        ("Структура", "У суммарных строк пусты «Длительность» и «% завершения»",
         all(r["% завершения"] == "" for r in sums)),
        ("Структура", "У суммарных строк проставлены даты, свёрнутые из потомков (DEC-25)",
         all(r["Начало"] and r["Окончание"] for r in sums)),
        ("Структура", "Ид. сквозной и совпадает с номером строки",
         all(r["Ид."] == i + 1 for i, r in enumerate(rows))),
        ("Связи", "Нотация русская (ОН · НН · ОО · НО), латиницы нет",
         not any(x in r["Предшественники"] for r in rows for x in ("FS", "SS", "FF", "SF"))),
        ("Связи", "«Последователи» — только Ид., без кода связи и лага (DEC-27)",
         not any(re.search(r"(ОН|НН|ОО|НО)|дн", r["Последователи"]) for r in rows)),
        ("Связи", "Лаги численные", not any("≥" in r["Предшественники"] for r in rows)),
        ("Длительности", "Все длительности в календарных днях",
         all(r["Длительность"].endswith("дней") for r in leaves)),
        ("Критический путь", f"Рассчитан ({crit} задач)", crit > 0),
        ("Сценарии", "Исключены по DEC-16 — не проверяются", True),
    ]
    sheet("Чек-лист", ["Раздел", "Пункт CLAUDE.md §9", "Результат"],
          [[a, b, "OK" if ok else "НАРУШЕНО"] for a, b, ok in checks], [20, 62, 14])

    wsc = wb.create_sheet("Сценарии")
    wsc["A1"] = "Блок не рассчитан"
    wsc["A1"].font = Font(bold=True)
    wsc["A2"] = ("DEC-16: решение владельца — правила сценариев поступят отдельной инструкцией "
                 "(standards.md §18). До неё три параметрических сценария и вероятностный расчёт "
                 "не строятся, выдача графика по этой причине не блокируется (CLAUDE.md §8).")
    wsc.column_dimensions["A"].width = 140
    wsc["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return [c for c in checks if not c[2]]


# ======================================================================
def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 2
    spec = Path(sys.argv[1])
    if not spec.is_absolute():
        spec = ROOT / spec
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / "out" / "ГРП.xlsx"
    if not out.is_absolute():
        out = ROOT / out

    project = json.loads(spec.read_text(encoding="utf-8"))
    b = Build(project)

    print(f"Проект: {project['название']}")
    print(f"Корпусов: {len(project['корпуса'])} · старт {project['старт_проекта']}\n")

    b.load_skeleton()
    b.check_stages()
    b.repair_defects()
    b.inherit_summary_links()
    b.apply_site_conditions()
    # Порядок существен: корпуса генерируются ПЕРВЫМИ. configure_parking при
    # отсутствии паркинга вставляет нулевой цикл внутрь блока К1, а
    # configure_corpuses заменяет весь участок корпусов целиком — при обратном
    # порядке вставленные строки были бы затёрты, а связи вех нулевого цикла
    # потерялись бы молча (CLAUDE.md §0.3).
    b.configure_corpuses()
    b.configure_zero_cycle()
    b.configure_parking()
    b.apply_standards()
    for corpus in project["корпуса"]:
        b.rebuild_monolith(corpus)

    nodes = b.schedule()          # проход 1
    b.thermal(nodes)
    b.wire_zos()
    nodes = b.schedule()          # проход 2 — после теплового блока
    b.report_anchors(nodes)
    rows = b.finalize(nodes)

    failed = write_excel(out, rows, b)

    finish = max(n.finish for n in nodes.values() if n.finish)
    crit = [n for n in nodes.values() if n.critical]
    names = {r["key"]: r["name"] for r in b.rows}
    print(f"Строк: {len(rows)} · вех: {sum(1 for r in rows if r['Длительность'] == '0 дней')}")
    print(f"Критический путь: {len(crit)} задач · финиш проекта {dfmt(finish)}")
    print(f"Обоснований: {len(b.rationale)} · допущений: {len(b.assumptions)}")
    if failed:
        print("\nЧек-лист НАРУШЕН:")
        for _, name, _ in failed:
            print(f"  ! {name}")
    if b.notices:
        print("\nУведомления:")
        for n in b.notices:
            print(f"  ! {n}")
    print(f"\nЗаписано: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
