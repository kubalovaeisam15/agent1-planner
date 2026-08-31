from pathlib import Path

import pytest

import build_grp
import mcp_server
import output_paths


@pytest.fixture
def output_env(tmp_path, monkeypatch):
    for key in ("AGENT1_RUNTIME", "AGENT1_DESKTOP_DIR", "XDG_DESKTOP_DIR", "CI",
                "USERPROFILE", "HOME"):
        monkeypatch.delenv(key, raising=False)
    monkeypatch.setattr(output_paths, "_windows_desktop", lambda: None)
    profile = tmp_path / "another-user"
    desktop = profile / "Desktop"
    desktop.mkdir(parents=True)
    repo = tmp_path / "repo"
    repo.mkdir()
    monkeypatch.setenv("USERPROFILE", str(profile))
    monkeypatch.setattr(mcp_server, "ROOT", repo)
    return repo, desktop


def test_profile_desktop_is_independent_of_cwd(output_env, tmp_path, monkeypatch):
    repo, desktop = output_env
    monkeypatch.chdir(tmp_path)
    assert output_paths.get_output_dir(repo) == desktop


def test_redirected_windows_desktop(output_env, tmp_path, monkeypatch):
    repo, _ = output_env
    redirected = tmp_path / "OneDrive" / "Рабочий стол"
    redirected.mkdir(parents=True)
    monkeypatch.setattr(output_paths, "_windows_desktop", lambda: str(redirected))
    assert output_paths.get_output_dir(repo) == redirected


def test_home_fallback(output_env, monkeypatch):
    repo, desktop = output_env
    monkeypatch.delenv("USERPROFILE")
    monkeypatch.setenv("HOME", str(desktop.parent))
    assert output_paths.get_output_dir(repo) == desktop


@pytest.mark.parametrize("setting", ["AGENT1_DESKTOP_DIR", "XDG_DESKTOP_DIR"])
def test_expanded_desktop_override(output_env, tmp_path, monkeypatch, setting):
    repo, desktop = output_env
    monkeypatch.setenv(setting, "$USERPROFILE/Desktop")
    assert output_paths.get_output_dir(repo) == desktop


@pytest.mark.parametrize("mode,ci", [("cloud", "false"), ("auto", "true")])
def test_cloud_ignores_desktop_and_creates_fallback(output_env, monkeypatch, mode, ci):
    repo, _ = output_env
    monkeypatch.setenv("AGENT1_RUNTIME", mode)
    monkeypatch.setenv("CI", ci)
    monkeypatch.setenv("AGENT1_DESKTOP_DIR", "bad-relative-path")
    expected = repo / "artifacts" / "output"
    assert output_paths.get_output_dir(repo, create=False) == expected
    assert not expected.exists()
    assert output_paths.get_output_dir(repo) == expected
    assert expected.is_dir()


def test_missing_desktop_auto_fallback_and_local_error(output_env):
    repo, desktop = output_env
    desktop.rmdir()
    assert output_paths.get_output_dir(repo) == repo / "artifacts" / "output"
    with pytest.raises(FileNotFoundError, match="Desktop"):
        output_paths.get_output_dir(repo, runtime="local")


def test_explicit_local_overrides_ci(output_env, monkeypatch):
    repo, desktop = output_env
    monkeypatch.setenv("CI", "true")
    assert output_paths.get_output_dir(repo, runtime="local") == desktop


@pytest.mark.parametrize("value", ["relative/Desktop", "$UNDEFINED_AGENT1_PATH/Desktop"])
def test_invalid_desktop_override_is_rejected(output_env, monkeypatch, value):
    repo, _ = output_env
    monkeypatch.setenv("AGENT1_DESKTOP_DIR", value)
    with pytest.raises(ValueError, match="absolute"):
        output_paths.get_output_dir(repo)


def test_invalid_runtime_and_root_directory_are_rejected(output_env, monkeypatch):
    repo, desktop = output_env
    with pytest.raises(ValueError, match="AGENT1_RUNTIME"):
        output_paths.get_output_dir(repo, runtime="clodu")
    monkeypatch.setenv("AGENT1_DESKTOP_DIR", desktop.anchor)
    with pytest.raises(ValueError, match="filesystem root"):
        output_paths.get_output_dir(repo)


def test_mcp_reads_output_but_keeps_inputs_and_traversal_restricted(output_env):
    repo, desktop = output_env
    ir = desktop / "schedule.json"
    ir.write_text("{}", encoding="utf-8")
    assert mcp_server._path(str(ir), allow_output=True) == ir
    with pytest.raises(mcp_server.ToolError, match="вне корня"):
        mcp_server._path(str(ir))  # inputs/templates remain repository-only
    with pytest.raises(mcp_server.ToolError, match="уже существует"):
        mcp_server._path(str(ir), allow_output=True, exists=False)
    for escaped in (desktop / ".." / "outside.json", repo / ".." / "outside.json",
                    desktop.parent / "Desktop-other" / "outside.json"):
        with pytest.raises(mcp_server.ToolError, match="вне корня"):
            mcp_server._path(str(escaped), allow_output=True, exists=False)


def test_mcp_rejects_symlink_escape(output_env, tmp_path):
    repo, desktop = output_env
    outside = tmp_path / "outside"
    outside.mkdir()
    link = desktop / "link"
    try:
        link.symlink_to(outside, target_is_directory=True)
    except OSError:
        pytest.skip("Creating symlinks requires OS permission")
    with pytest.raises(mcp_server.ToolError, match="вне корня"):
        mcp_server._path(str(link / "file.json"), allow_output=True, exists=False)
    (repo / "artifacts").symlink_to(outside, target_is_directory=True)
    with pytest.raises(ValueError, match="outside"):
        output_paths.get_output_dir(repo, runtime="cloud")


def test_defaults_are_unique_without_creating_files(output_env, monkeypatch):
    repo, desktop = output_env
    monkeypatch.setattr(output_paths, "ROOT", repo)
    first = output_paths.new_output_path(".xlsx")
    second = output_paths.new_output_path(".xlsx")
    assert first.parent == second.parent == desktop
    assert first != second
    assert not first.exists() and not second.exists()


@pytest.mark.parametrize("runtime", ["local", "cloud"])
def test_mcp_default_build_and_readback(output_env, monkeypatch, runtime):
    from test_mcp_server import sample_schedule

    repo, desktop = output_env
    monkeypatch.setenv("AGENT1_RUNTIME", runtime)
    monkeypatch.setattr(mcp_server, "_require_context_ready", lambda: None)
    spec = repo / "input.json"
    spec.write_text("{}", encoding="utf-8")

    def fake_build(argv):
        Path(argv[1]).write_bytes(b"xlsx")
        Path(argv[3]).write_text(sample_schedule().to_json(), encoding="utf-8")
        return 0

    monkeypatch.setattr(build_grp, "main", fake_build)
    result = mcp_server.call_tool("schedule_build", {"spec_path": "input.json"})
    assert not result["isError"], result
    paths = result["structuredContent"]
    expected = desktop if runtime == "local" else repo / "artifacts" / "output"
    assert Path(paths["xlsx_path"]).parent == expected
    assert Path(paths["ir_path"]).parent == expected
    summary = mcp_server.call_tool("schedule_summary", {"ir_path": paths["ir_path"]})
    assert not summary["isError"], summary


def test_cli_refuses_existing_excel_or_ir_before_loading_inputs(output_env):
    _, desktop = output_env
    existing = desktop / "keep.xlsx"
    existing.write_bytes(b"keep")
    assert build_grp.main(["missing.json", str(existing)]) == 1
    assert build_grp.main(["missing.json", str(desktop / "new.xlsx"),
                           "--ir", str(existing)]) == 1
    assert existing.read_bytes() == b"keep"


def test_permission_error_is_not_silently_redirected(output_env, monkeypatch):
    repo, _ = output_env
    def deny(*args, **kwargs):
        raise PermissionError("denied")
    monkeypatch.setattr(Path, "mkdir", deny)
    with pytest.raises(PermissionError):
        output_paths.get_output_dir(repo)


@pytest.mark.parametrize("runtime", ["local", "cloud"])
def test_cli_builds_real_excel_and_ir_in_default_directory(output_env, monkeypatch, runtime):
    from openpyxl import load_workbook
    from schedule_ir import ScheduleProject, validate_schedule_ir

    spec = build_grp.ROOT / "tests" / "etalon_project.json"
    repo, desktop = output_env
    monkeypatch.setattr(build_grp, "ROOT", repo)
    monkeypatch.setenv("AGENT1_RUNTIME", runtime)
    assert build_grp.main([str(spec)]) == 0
    expected = desktop if runtime == "local" else repo / "artifacts" / "output"
    xlsx, = expected.glob("*.xlsx")
    ir, = expected.glob("*.ir.json")
    workbook = load_workbook(xlsx, read_only=True)
    try:
        assert "ГРП" in workbook.sheetnames
    finally:
        workbook.close()
    assert not validate_schedule_ir(ScheduleProject.from_json(ir.read_text(encoding="utf-8")))


def test_mcp_mpp_and_report_paths_support_desktop(output_env, monkeypatch):
    from test_mcp_server import sample_schedule

    repo, desktop = output_env
    ir = desktop / "schedule.ir.json"
    ir.write_text(sample_schedule().to_json(), encoding="utf-8")
    monkeypatch.setattr(mcp_server, "_require_context_ready", lambda: None)

    def fake_export(ir_path, mpp_path, **kwargs):
        assert ir_path == ir
        assert mpp_path.parent == desktop
        mpp_path.write_bytes(b"mpp")
        return {}

    monkeypatch.setattr(mcp_server, "export_mpp", fake_export)
    exported = mcp_server.call_tool("mpp_export", {"ir_path": str(ir)})
    assert not exported["isError"], exported
    mpp_path = exported["structuredContent"]["mpp_path"]
    report = desktop / "report.json"
    monkeypatch.setattr(mcp_server, "read_mpp", lambda *a, **kw: object())
    monkeypatch.setattr(mcp_server, "validate_snapshot", lambda *a: [])
    monkeypatch.setattr(mcp_server, "compare_with_ir", lambda *a: [])
    monkeypatch.setattr(mcp_server, "report_dict", lambda *a: {
        "project": {}, "result": {}, "issues": []})
    arguments = {"mpp_path": mpp_path, "ir_path": str(ir), "report_path": str(report)}
    validated = mcp_server.call_tool("mpp_validate", arguments)
    assert not validated["isError"], validated
    assert report.is_file()
    assert mcp_server.call_tool("mpp_validate", arguments)["isError"]
