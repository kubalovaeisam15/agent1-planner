# -*- coding: utf-8 -*-
"""DEC-32: две карты вех — фаза A и фаза B — являются выборкой ГРП."""
from __future__ import annotations

import json
from pathlib import Path
from uuid import uuid4

import openpyxl

import build_grp
from validate_grp import load, load_milestone_maps, load_reserve_base, validate

ROOT = Path(__file__).resolve().parents[1]


def build_workbook(path: Path) -> None:
    project = json.loads((ROOT / "tests" / "etalon_project.json").read_text(encoding="utf-8"))
    b = build_grp.Build(project)
    b.load_skeleton()
    b.check_stages()
    b.repair_defects()
    b.inherit_summary_links()
    b.apply_site_conditions()
    b.configure_corpuses()
    b.configure_zero_cycle()
    b.configure_parking()
    b.apply_finishing_scope()
    b.apply_standards()
    for corpus in project["корпуса"]:
        b.rebuild_monolith(corpus)
    nodes = b.schedule()
    b.thermal(nodes)
    b.wire_zos()
    nodes = b.schedule()
    rows = b.finalize(nodes)
    build_grp.write_excel(path, rows, b)


def test_two_maps_exist_and_match_grp():
    # Системный TEMP на некоторых рабочих станциях закрыт политиками доступа.
    # Используем уже существующую и игнорируемую Git папку проекта.
    out_dir = ROOT / "out"
    out_dir.mkdir(exist_ok=True)
    path = out_dir / f"test-dec32-{uuid4().hex}.xlsx"
    try:
        build_workbook(path)
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            assert "Карта вех A" in wb.sheetnames
            assert "Карта вех B" in wb.sheetnames
        finally:
            wb.close()

        maps = load_milestone_maps(path)
        assert maps["Карта вех A"]
        assert maps["Карта вех B"]
        assert all(not row["Дата вехи утверждённая"] for row in maps["Карта вех A"])
        assert all(row["Дата вехи утверждённая"] for row in maps["Карта вех B"])
        assert validate(load(path), load_reserve_base(path), maps) == 0
    finally:
        path.unlink(missing_ok=True)
