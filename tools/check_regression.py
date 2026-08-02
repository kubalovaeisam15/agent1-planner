# -*- coding: utf-8 -*-
"""Регрессионная проверка по критериям приёмки CLAUDE.md §10.

    python tools/check_regression.py out/ГРП_эталон.xlsx

Сверяет выдачу с tests/etalon_expected.md §1–§2: даты РС и РВЭ, длительности
двух фаз по отдельности и покорпусные ориентиры. Сквозная длительность не
проверяется — DEC-19.

Адресация вех — по наименованию: колонки СДР в шаблоне v2 нет (typGRP.md §2).

Код возврата: 0 — все жёсткие критерии в допуске; 1 — есть выход за ±14 дн.
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# Консоль Windows может быть в cp1251: не даём выводу падать на символах,
# которых нет в её кодировке (стрелки, типографика).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")

ROOT = Path(__file__).resolve().parents[1]
TOL = 14  # ±2 недели — CLAUDE.md §10

# Эталон — tests/etalon_expected.md. Правка «под результат» запрещена.
HARD = {
    "РС": ("Разрешение на строительство (РС) получено", "02.04.2024"),
    "РВЭ": ("Получено РВЭ", "31.12.2026"),
}
PHASES = {
    "Фаза A (МЗ → РС)": ("Маркетинговое задание утверждено",
                         "Разрешение на строительство (РС) получено", 447),
    "Фаза РС → РВЭ": ("Разрешение на строительство (РС) получено",
                      "Получено РВЭ", 1003),
}
# Ориентиры второго уровня — etalon_expected.md §2. Покорпусные вехи проверяют
# главное: одно правило должно давать РАЗНЫЕ числа для К1 (26 эт.) и К2 (55 эт.).
LANDMARKS = {
    "Завершено свайное поле": ("Завершено свайное поле", "04.07.2024"),
    "ЯКОРЬ К1 (кровля/парапет)": ("К1. Кровля/Парапет Монолит", "04.05.2025"),
    "ЯКОРЬ К2 (кровля/парапет)": ("К2. Кровля/Парапет Монолит", "25.10.2025"),
    "TC_perm К1": ("К1. Закрыт тепловой контур по корпусу", "23.07.2025"),
    "TC_perm К2": ("К2. Закрыт тепловой контур по корпусу", "03.03.2026"),
    "Пуск тепла К1": ("К1. Пуск тепла корпус", "04.11.2025"),
    "Пуск тепла К2": ("К2. Пуск тепла корпус", "20.12.2025"),
    "Пуск тепла в паркинге": ("П1. Пуск тепла в паркинге", "06.10.2025"),
    "Финиш отделки К1": ("К1. По договору Вестибюль", "08.05.2026"),
    "Финиш отделки К2": ("К2. По договору Вестибюль", "20.10.2026"),
    "Получен ЗОС": ("Получен ЗОС", "26.12.2026"),
    "Переданы квартиры с отделкой": ("Переданы квартиры с отделкой", "01.10.2027"),
}
# Длительности, выводимые правилом, а не копируемые из шаблона (etalon_expected §2.2).
DURATIONS = {
    "Монолит К1 (26 эт.)": ("К1. Монтаж монолитных конструкций выше отм. 0.000", 195),
    "Монолит К2 (55 эт.)": ("К2. Монтаж монолитных конструкций выше отм. 0.000", 369),
}


def d(s: str):
    return datetime.strptime(s, "%d.%m.%Y").date()


def load(path: Path) -> dict[str, dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c else "" for c in next(rows)]
    out: dict[str, dict] = {}
    for raw in rows:
        rec = {h: ("" if v is None else str(v).strip()) for h, v in zip(header, raw)}
        name = rec.get("Название задачи", "")
        if name and name not in out:      # первое вхождение — сама веха, не её копия
            out[name] = rec
    return out


def finish_of(tasks: dict, name: str) -> str | None:
    r = tasks.get(name)
    return (r.get("Окончание") or None) if r else None


def span_of(tasks: dict, name: str) -> int | None:
    r = tasks.get(name)
    if not r or not r.get("Начало") or not r.get("Окончание"):
        return None
    return (d(r["Окончание"]) - d(r["Начало"])).days


def main() -> int:
    if len(sys.argv) != 2:
        print(__doc__)
        return 2
    path = Path(sys.argv[1])
    if not path.is_absolute():
        path = ROOT / path
    tasks = load(path)

    print(f"=== Регрессия: {path.name} (CLAUDE.md §10, допуск ±{TOL} дн) ===\n")
    failed = 0

    print("-- Жёсткие критерии --")
    for name, (task, expected) in HARD.items():
        got = finish_of(tasks, task)
        if not got:
            print(f"  НЕТ ДАННЫХ {name}: «{task}» отсутствует или без даты")
            failed += 1
            continue
        delta = (d(got) - d(expected)).days
        ok = abs(delta) <= TOL
        failed += 0 if ok else 1
        print(f"  {'OK  ' if ok else 'МИМО'} {name}: {got} против эталона {expected} "
              f"({delta:+d} дн)")

    print("\n-- Длительности фаз (DEC-19, пофазно) --")
    for name, (a_task, b_task, expected) in PHASES.items():
        a, b = finish_of(tasks, a_task), finish_of(tasks, b_task)
        if not a or not b:
            print(f"  НЕТ ДАННЫХ {name}")
            failed += 1
            continue
        got = (d(b) - d(a)).days
        delta = got - expected
        ok = abs(delta) <= TOL
        failed += 0 if ok else 1
        print(f"  {'OK  ' if ok else 'МИМО'} {name}: {got} дн против эталона {expected} дн "
              f"({delta:+d} дн)")

    print("\n-- Длительности, выведенные правилом (не скопированные) --")
    for name, (task, expected) in DURATIONS.items():
        got = span_of(tasks, task)
        if got is None:
            print(f"  —    {name}: нет данных")
            continue
        mark = "OK  " if got == expected else "!   "
        print(f"  {mark} {name}: {got} дн против {expected} дн ({got - expected:+d} дн)")

    print("\n-- Ориентиры второго уровня (не критерии приёмки) --")
    for name, (task, expected) in LANDMARKS.items():
        got = finish_of(tasks, task)
        if not got:
            print(f"  —    {name}: нет данных")
            continue
        delta = (d(got) - d(expected)).days
        mark = "OK  " if abs(delta) <= TOL else "!   "
        print(f"  {mark} {name}: {got} против {expected} ({delta:+d} дн)")

    print(f"\nИтог: нарушений жёстких критериев — {failed}")
    if failed:
        print("Разбирается причина в standards.md / bindings.md / входных данных.")
        print("Правка tests/etalon_expected.md «под результат» запрещена (CLAUDE.md §10).")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
